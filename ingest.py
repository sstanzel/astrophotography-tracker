#!/usr/bin/env python3
"""
ingest.py — populate the astrophotography tracker SQLite database.

Walks every configured capture library, parses every session folder and
FITS filename, walks the shared calibration library, and upserts everything into
the schema defined in schema.sql.

Idempotent: safe to re-run. Sessions upsert on their natural key so session_id
stays stable; a session's frames are deleted and re-inserted on each run so the
counts always reflect what's on disk right now.

Usage:
    python3 ingest.py [--db PATH] [--schema PATH] [--xlsx PATH] [--quiet]

Defaults assume the script lives in  _organization/tracker/  and
writes the database next to it. Run with no arguments after a capture night to
refresh the tracker.
"""

import os, re, sys, json, sqlite3, argparse, datetime

# --------------------------------------------------------------------------
# Library locations come from config.toml (via astro_config) — not hardcoded.
# --------------------------------------------------------------------------
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import astro_config  # noqa: E402

# Top-level folders that are "other captures" — recorded but exempt from
# deep-sky integration totals and v2 naming enforcement. Deliberately one
# level deep: the science topics are top-level buckets, not Science/{topic}.
# "Moon Daytime"/"Moon Nighttime" merge into "Moon" and the time-lapse bucket
# folds into "Nightscapes" — the old names stay recognized until those two
# library cleanups are done (work items in Things, 2026-07-10).
OTHER_CAPTURE_FOLDERS = {
    "ASI EAA",
    "Asteroids Comets",
    "As_misc",
    "Moon",
    "Moon Daytime",
    "Moon Nighttime",
    "Planets",
    "Nightscapes",
    "As_Tl_Astrophotography Timelapse",
    "Astrometry",
    "Spectroscopy",
    "Photometry",
    "Exoplanets",
    "Double Stars",
}

# Folders to skip entirely at the target level.
SKIP_TOPLEVEL = {"_organization", "_Calibration Library", "_sessions to organize"}

# Parser import — fits_parser.py lives in the same directory.
from fits_parser import parse as parse_fits, frame_kind, safe, is_non_science  # noqa: E402

SESSION_RE = re.compile(
    r"^(?P<target>[A-Za-z0-9_+\-]+)\s+(?P<scope>\S+)\s+(?P<sensor>\S+)\s+(?P<date>\d{4}-\d{2}-\d{2})$"
)
CATALOG_RE = re.compile(r"^(?P<cat>M|NGC|NCG|IC|C|SH2|LDN|HR)\s+(?P<rest>.+)$")
DATE_RE = re.compile(r"(\d{4}-\d{2}-\d{2})")


# ==========================================================================
# Helpers
# ==========================================================================
def parse_target_folder(name):
    """Return dict(target_id, catalog, number, common_name, is_other)."""
    if name in OTHER_CAPTURE_FOLDERS:
        return dict(
            target_id=name.replace(" ", "_"),
            catalog="Other",
            number=None,
            common_name=name,
            is_other=1,
        )
    m = CATALOG_RE.match(name)
    if not m:
        # Named-star or non-catalog folder (e.g. "Capella star")
        return dict(
            target_id=name.replace(" ", "_"), catalog="?", number=None, common_name=name, is_other=0
        )
    cat = m.group("cat")
    tokens = m.group("rest").split()
    num_tokens = []
    while tokens and tokens[0].isdigit():
        num_tokens.append(tokens.pop(0))
    number = " ".join(num_tokens) if num_tokens else None
    common = " ".join(tokens) if tokens else None
    first_num = num_tokens[0] if num_tokens else (m.group("rest").split()[0])
    return dict(
        target_id=f"{cat}_{first_num}", catalog=cat, number=number, common_name=common, is_other=0
    )


def walk_fits(session_path):
    """Yield (abs_path, is_rejected, parse_match) for every FITS-extension file
    under a session. parse_match is None when the filename did not parse — the
    caller counts those instead of silently dropping them."""
    for root, dirs, files in os.walk(session_path):
        is_rej = "/Rejected" in root or root.endswith("/Rejected")
        for f in files:
            if f == ".DS_Store" or f.startswith("._"):
                continue
            if not f.lower().endswith((".fit", ".fits", ".xisf")):
                continue
            yield os.path.join(root, f), is_rej, parse_fits(f)


def count_tree(path):
    """Return (file_count, total_bytes) recursively, ignoring macOS noise."""
    fc, sz = 0, 0
    for root, dirs, files in os.walk(path):
        for f in files:
            if f == ".DS_Store" or f.startswith("._"):
                continue
            try:
                sz += os.path.getsize(os.path.join(root, f))
            except OSError:
                pass
            fc += 1
    return fc, sz


# --------------------------------------------------------------------------
# Validation-input readers: notes.toml, locations.toml, FITS headers
# --------------------------------------------------------------------------
def load_locations(path):
    """Minimal TOML-subset parse of locations.toml.
    Returns {site_name: {'lat':float, 'lon':float, 'bortle':value}}."""
    sites, cur = {}, None
    if not path or not os.path.exists(path):
        return {}
    with open(path, encoding="utf-8") as fh:
        for line in fh:
            s = line.strip()
            if not s or s.startswith("#"):
                continue
            mh = re.match(r'\[\s*"?(.+?)"?\s*\]\s*$', s)
            if mh:
                cur = mh.group(1)
                sites[cur] = {}
                continue
            if cur is None:
                continue
            mk = re.match(r"(\w+)\s*=\s*(.+?)\s*(?:#.*)?$", s)
            if mk:
                k, raw = mk.group(1), mk.group(2).strip()
                if raw[:1] in "\"'":
                    sites[cur][k] = raw.strip("\"'")
                else:
                    try:
                        sites[cur][k] = float(raw)
                    except ValueError:
                        sites[cur][k] = raw
    out = {}
    for name, d in sites.items():
        if "latitude" in d and "longitude" in d:
            out[name] = {
                "lat": float(d["latitude"]),
                "lon": float(d["longitude"]),
                "bortle": d.get("bortle"),
            }
    return out


def parse_toml_tables(text, name):
    """Parse `[[name]]` array-of-tables blocks into a list of dicts.

    Zero-dependency: string, number and boolean scalar values only. A block ends
    at the next section header of any kind.

    Args:
        text: TOML text.
        name: the array-of-tables name (e.g. 'published').

    Returns:
        List of dicts, one per [[name]] block, in file order.
    """
    out, cur = [], None
    for raw in text.splitlines():
        line = raw.strip()
        if not line or line.startswith("#"):
            continue
        if line == f"[[{name}]]":
            cur = {}
            out.append(cur)
            continue
        if line.startswith("["):
            cur = None
            continue
        if cur is None:
            continue
        ms = re.match(r'(\w+)\s*=\s*"(.*?)"', line)
        mb = re.match(r"(\w+)\s*=\s*(true|false)\s*$", line, re.I)
        mn = re.match(r"(\w+)\s*=\s*(-?[\d.]+)\s*$", line)
        if ms:
            cur[ms.group(1)] = ms.group(2)
        elif mb:
            cur[mb.group(1)] = mb.group(2).lower() == "true"
        elif mn:
            cur[mn.group(1)] = mn.group(2)
    return out


def read_notes_toml(session_path, session_name):
    """Parse a session's `{name} notes.toml`. Regex parse — zero dependencies.

    Returns a dict with sky metadata, the `edited` flag, an explicit `culled`
    flag (for a reviewed-kept-all session), and the `[[published]]`/`[[printed]]`
    entry lists (each a list of dicts).
    """
    out = dict(
        present=False,
        location=None,
        moon_phase=None,
        moon_illumination=None,
        moon_age_days=None,
        edited=False,
        culled=False,
        flats_with=None,
        published=[],
        printed=[],
        todos=[],
    )
    p = os.path.join(session_path, f"{session_name} notes.toml")
    if not os.path.isfile(p):
        return out
    out["present"] = True
    try:
        txt = open(p, encoding="utf-8").read()
    except OSError:
        return out
    for flag in ("edited", "culled"):
        fm = re.search(r"^\s*" + flag + r"\s*=\s*(true|false)", txt, re.M | re.I)
        out[flag] = bool(fm) and fm.group(1).lower() == "true"
    out["published"] = parse_toml_tables(txt, "published")
    out["printed"] = parse_toml_tables(txt, "printed")
    m = re.search(r'^location\s*=\s*"([^"]*)"', txt, re.M)
    if m and m.group(1):
        out["location"] = m.group(1)
    m = re.search(r'^moon_phase\s*=\s*"([^"]*)"', txt, re.M)
    if m and m.group(1):
        out["moon_phase"] = m.group(1)
    m = re.search(r"^moon_illumination\s*=\s*([0-9.]+)", txt, re.M)
    if m:
        out["moon_illumination"] = float(m.group(1))
    m = re.search(r"^moon_age_days\s*=\s*([0-9.]+)", txt, re.M)
    if m:
        out["moon_age_days"] = float(m.group(1))
    # [calibration] flats = "<sibling session folder>" — shared-flat night
    # pointer: this session's flats live in that sibling's folder.
    m = re.search(r'^flats\s*=\s*"([^"]*)"', txt, re.M)
    if m and m.group(1):
        out["flats_with"] = m.group(1)
    # [future_processing] todo = [ "…", "…" ]  — reprocessing to-do items.
    am = re.search(r"(?ms)^\s*todo\s*=\s*\[(.*?)\]", txt)
    out["todos"] = re.findall(r'"([^"]*)"', am.group(1)) if am else []
    return out


def insert_publications(cur, target_id, session_id, integration_id, published, printed):
    """Rebuild the publications rows for one session or integration.

    Deletes any existing rows for that session/integration, then inserts one row
    per [[published]] and [[printed]] entry. Source of truth is the toml lists.

    Args:
        cur: DB cursor.
        target_id: the target the image belongs to.
        session_id: session id, or None for an integration.
        integration_id: integration id, or None for a session.
        published: list of dicts (kind/url/title/date/note).
        printed: list of dicts (title/date/note).
    """
    if session_id is not None:
        cur.execute("DELETE FROM publications WHERE session_id=?", (session_id,))
    if integration_id is not None:
        cur.execute("DELETE FROM publications WHERE integration_id=?", (integration_id,))
    for e in published:
        k = (e.get("kind") or "other").lower()
        if k not in ("astrobin", "social", "other"):
            k = "other"
        cur.execute(
            """INSERT INTO publications(target_id, session_id, integration_id,
                       kind, url, title, published_at, notes)
                       VALUES(?,?,?,?,?,?,?,?)""",
            (
                target_id,
                session_id,
                integration_id,
                k,
                e.get("url"),
                e.get("title"),
                e.get("date"),
                e.get("note"),
            ),
        )
    for e in printed:
        cur.execute(
            """INSERT INTO publications(target_id, session_id, integration_id,
                       kind, url, title, published_at, notes)
                       VALUES(?,?,?,'print',?,?,?,?)""",
            (
                target_id,
                session_id,
                integration_id,
                e.get("url"),
                e.get("title"),
                e.get("date"),
                e.get("note"),
            ),
        )


def detect_method(container_path):
    """Return how a session/integration was stacked, from its working folders.

    'PI Magic' or 'PixInsight' when the matching working folder still holds
    files (PI Magic wins if both do); 'other' when the Results folder has files
    but the working folders are empty (e.g. already cleaned); None when nothing
    indicates it was integrated. Callers persist a specific method so it survives
    a later cleanup.

    Args:
        container_path: absolute path of the session or integration folder.

    Returns:
        'PI Magic' | 'PixInsight' | 'other' | None.
    """

    def has_files(sub):
        p = os.path.join(container_path, sub)
        if not os.path.isdir(p):
            return False
        for _r, _d, files in os.walk(p):
            if any(f != ".DS_Store" and not f.startswith("._") for f in files):
                return True
        return False

    if has_files("PI Magic"):
        return "PI Magic"
    if has_files("PI Process"):
        return "PixInsight"
    if has_files(f"{os.path.basename(container_path)} Results"):
        return "other"
    return None


def span_matches(date, span):
    """True if a YYYY-MM-DD date falls in a span: None/'all', a year, or a range.

    Args:
        date: 'YYYY-MM-DD' capture date.
        span: 'all'/None, a year '2026', or a range '2024-2026'.

    Returns:
        Whether the date's year is covered by the span.
    """
    if not span or span == "all":
        return True
    year = date[:4]
    m = re.fullmatch(r"(\d{4})-(\d{4})", span)
    if m:
        return m.group(1) <= year <= m.group(2)
    return year == span


def resolve_auto_members(target_path, rig, span, exclude):
    """Return session folder names under a target that match a rig+span rule.

    Args:
        target_path: absolute path of the target folder.
        rig: '<scope> <sensor>' to require one rig, or falsy for any rig (composite).
        span: span filter passed to span_matches().
        exclude: iterable of session folder names to drop.

    Returns:
        Sorted list of matching session folder names.
    """
    excl = set(exclude or [])
    scope_want = sensor_want = None
    if rig:
        parts = rig.split()
        if len(parts) >= 2:
            scope_want, sensor_want = parts[0], parts[1]
    out = []
    for sname in sorted(os.listdir(target_path)):
        if sname == "integrations" or sname.startswith((".", "_")):
            continue
        if not os.path.isdir(os.path.join(target_path, sname)):
            continue
        m = SESSION_RE.match(sname)
        if not m or not span_matches(m.group("date"), span) or sname in excl:
            continue
        if scope_want and (m.group("scope") != scope_want or m.group("sensor") != sensor_want):
            continue
        out.append(sname)
    return out


def read_integration_toml(path):
    """Parse an integration.toml manifest. Returns a dict, or None if absent.

    Handles the rule-based format ([membership] rig/span + [built] sessions) and
    the legacy explicit-members format (top-level `members` list). The parse is a
    flat, section-agnostic regex — keys are matched wherever they appear."""
    if not os.path.isfile(path):
        return None
    try:
        txt = open(path, encoding="utf-8").read()
    except OSError:
        return None

    def s(key):
        m = re.search(r"^\s*" + key + r'\s*=\s*"([^"]*)"', txt, re.M)
        return m.group(1) if m else None

    def i(key):
        m = re.search(r"^\s*" + key + r"\s*=\s*(\d+)", txt, re.M)
        return int(m.group(1)) if m else None

    def flt(key):
        m = re.search(r"^\s*" + key + r"\s*=\s*([\d.]+)", txt, re.M)
        return float(m.group(1)) if m else None

    def b(key):
        m = re.search(r"^\s*" + key + r"\s*=\s*(true|false)", txt, re.M | re.I)
        return bool(m) and m.group(1).lower() == "true"

    def arr(key):
        m = re.search(r"(?ms)^\s*" + key + r"\s*=\s*\[(.*?)\]", txt)
        return re.findall(r'"([^"]*)"', m.group(1)) if m else []

    mode = s("mode")
    return {
        "mode": mode if mode in ("auto", "pinned") else None,
        "rig": s("rig"),
        "span": s("span"),
        "goal_hours": flt("goal_hours"),
        "exclude": arr("exclude"),
        "members": arr("members"),  # legacy explicit list
        "built_sessions": arr("sessions"),  # [built].sessions
        "kind": s("kind"),
        "version": i("version") or 1,
        "edited": b("edited"),
        "published": parse_toml_tables(txt, "published"),
        "printed": parse_toml_tables(txt, "printed"),
    }


def read_fits_header(path):
    """Pull a few keys from a FITS primary header (first 28800 bytes).
    Returns dict(site_lat, site_lon, instrument, telescope); any may be None.
    Safe on unreadable files and on XISF (which simply won't match)."""
    out = dict(site_lat=None, site_lon=None, instrument=None, telescope=None)
    try:
        with open(path, "rb") as fh:
            blob = fh.read(28800).decode("latin-1", "ignore")
    except OSError:
        return out
    mn = re.search(r"SITELAT\s*=\s*([-+0-9.eE]+)", blob)
    if mn:
        out["site_lat"] = float(mn.group(1))
    mn = re.search(r"SITELONG\s*=\s*([-+0-9.eE]+)", blob)
    if mn:
        out["site_lon"] = float(mn.group(1))
    mt = re.search(r"INSTRUME\s*=\s*'([^']*)'", blob)
    if mt:
        out["instrument"] = mt.group(1).strip()
    mt = re.search(r"TELESCOP\s*=\s*'([^']*)'", blob)
    if mt:
        out["telescope"] = mt.group(1).strip()
    return out


# ==========================================================================
# Ingest steps
# ==========================================================================
SESSION_NEW_COLS = [
    ("notes_toml_present", "INTEGER DEFAULT 0"),
    ("unparsed_file_count", "INTEGER DEFAULT 0"),
    ("other_image_count", "INTEGER DEFAULT 0"),
    ("results_file_count", "INTEGER DEFAULT 0"),
    ("fits_site_lat", "REAL"),
    ("fits_site_lon", "REAL"),
    ("fits_instrument", "TEXT"),
    ("stage_culled", "INTEGER DEFAULT 0"),  # 2 Culled
    ("stage_integrate", "INTEGER DEFAULT 0"),  # 3 Integrated (single-session)
    ("stage_edit", "INTEGER DEFAULT 0"),  # 4 Edited
    ("stage_publish", "INTEGER DEFAULT 0"),  # 5 Published
    ("stage_print", "INTEGER DEFAULT 0"),  # 6 Printed
    ("integration_method", "TEXT"),  # PixInsight|PI Magic|other
    ("astrobin_url", "TEXT"),
    ("flats_source", "TEXT"),  # here|with sibling|library|none (derived)
    ("flats_ref", "TEXT"),  # sibling session folder or library set path
    ("flats_note_ref", "TEXT"),  # notes.toml [calibration] flats pointer
]

# Columns added to the integrations table on an existing DB (no-op on a fresh
# one). The CHECK on membership_mode lives in schema.sql for new DBs only.
INTEGRATION_NEW_COLS = [
    ("membership_mode", "TEXT NOT NULL DEFAULT 'auto'"),
    ("goal_hours", "REAL"),
    ("integration_method", "TEXT"),
]

# frames table rebuild — used by apply_migrations when an old DB's grammar CHECK
# predates 'asiair_dslr'. SQLite can't ALTER a CHECK; the frames table holds zero
# manual data (rebuilt every ingest), so dropping + recreating it is safe.
FRAMES_TABLE_REBUILD = """
DROP TABLE IF EXISTS frames;
CREATE TABLE frames (
    frame_id            INTEGER PRIMARY KEY AUTOINCREMENT,
    session_id          INTEGER NOT NULL REFERENCES sessions(session_id) ON DELETE CASCADE,
    frame_type          TEXT NOT NULL CHECK (frame_type IN ('light','flat','dark','dark_flat','bias')),
    is_rejected         INTEGER NOT NULL DEFAULT 0,
    exp_value           REAL NOT NULL,
    exp_unit            TEXT NOT NULL CHECK (exp_unit IN ('s','ms')),
    exp_s               REAL NOT NULL,
    binning             TEXT NOT NULL,
    camera_short        TEXT NOT NULL,
    gain                INTEGER,
    temp_c              REAL,
    rotation_deg        REAL,
    captured_at_utc     DATETIME NOT NULL,
    filter              TEXT,
    hfr                 REAL,
    rms_arcsec          REAL,
    grammar             TEXT NOT NULL CHECK (grammar IN
                          ('asiair_sci','asiair_cal','asiair_dslr',
                           'nina_legacy','nina_v2','nina_cal')),
    file_path           TEXT NOT NULL,
    file_size_bytes     INTEGER,
    sequence_index      INTEGER,
    UNIQUE (session_id, file_path)
);
CREATE INDEX idx_frames_session    ON frames(session_id);
CREATE INDEX idx_frames_type       ON frames(frame_type);
CREATE INDEX idx_frames_captured   ON frames(captured_at_utc);
CREATE INDEX idx_frames_qc_hfr_rms ON frames(hfr, rms_arcsec);
"""

# Non-FITS raw camera formats — a DSLR session holds these instead of .fit files.
RAW_IMAGE_EXT = (".cr3", ".cr2", ".nef", ".arw", ".raf", ".dng", ".orf", ".rw2")
# Subfolders that hold processing output, not raw captures.
PROCESSING_DIRS = ("PI Process", "PI Magic")


def in_processing_area(abs_path, session_path):
    """True if abs_path sits inside a processing-output subfolder of the session
    (PI Process, PI Magic, a .pxiproject bundle, or a Results folder)."""
    rel = os.path.relpath(abs_path, session_path)
    for part in rel.split(os.sep)[:-1]:  # directory components only
        if part in PROCESSING_DIRS or part.endswith(".pxiproject") or part.endswith(" Results"):
            return True
    return False


def has_master_file(path):
    """True if a produced calibration master (`master*.xisf/.fit`) sits under path.

    File-based — a set 'has a master' only when the master file is actually there,
    not because of where the folder lives. Drives the 'build masters' worklist.
    """
    for _r, _d, files in os.walk(path):
        for f in files:
            if f.lower().startswith("master") and f.lower().endswith((".xisf", ".fit", ".fits")):
                return True
    return False


def detect_set_gain(path):
    """Most common gain/ISO token in the names under a calibration set folder.

    Bias is gain/ISO-dependent but bias sets aren't gain-foldered: ASIAir frame
    names carry `gain###` and ASIAir DSLR subfolders carry `ISO###`, so read it
    from there. ISO lands in the same column lights use for it (frames.gain),
    keeping coverage matching apples-to-apples. None when no token is found
    (such a set matches any gain, like a dark set with no temp folder).
    """
    counts = {}
    for _r, dirs, files in os.walk(path):
        for name in dirs + files:
            for tok in re.split(r"[_\s]+", name):
                m = re.match(r"^[Gg]ain(-?\d+)$", tok) or re.match(r"^ISO(\d+)$", tok)
                if m:
                    v = int(m.group(1))
                    counts[v] = counts.get(v, 0) + 1
    return max(counts, key=counts.get) if counts else None


def count_raw_images(session_path):
    """Count non-FITS raw camera files under a session (DSLR sessions)."""
    n = 0
    for root, _dirs, files in os.walk(session_path):
        for f in files:
            if f.lower().endswith(RAW_IMAGE_EXT):
                n += 1
    return n


def init_db(db_path, schema_path):
    fresh = not os.path.exists(db_path)
    con = sqlite3.connect(db_path)
    con.execute("PRAGMA foreign_keys = ON")
    if fresh:
        con.executescript(open(schema_path).read())
        con.commit()
    return con


def apply_migrations(con):
    """Bring an existing tracker.db up to the current schema without data loss.
    Effectively a no-op on a freshly created DB. Lets the validation layer be
    added in place — manual stage flags and goals in an old DB are preserved."""
    cur = con.cursor()
    have = {r[1] for r in cur.execute("PRAGMA table_info(sessions)")}
    for col, decl in SESSION_NEW_COLS:
        if col not in have:
            cur.execute(f"ALTER TABLE sessions ADD COLUMN {col} {decl}")
    for tbl in ("scopes", "sensors"):
        cols = {r[1] for r in cur.execute(f"PRAGMA table_info({tbl})")}
        if "from_registry" not in cols:
            cur.execute(
                f"ALTER TABLE {tbl} ADD COLUMN " f"from_registry INTEGER NOT NULL DEFAULT 0"
            )
    cur.executescript("""
        CREATE TABLE IF NOT EXISTS validation_findings (
            finding_id  INTEGER PRIMARY KEY AUTOINCREMENT,
            severity    TEXT NOT NULL,
            code        TEXT NOT NULL,
            scope       TEXT NOT NULL,
            session_id  INTEGER REFERENCES sessions(session_id) ON DELETE CASCADE,
            ref_path    TEXT,
            message     TEXT NOT NULL,
            detected_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );
        CREATE INDEX IF NOT EXISTS idx_findings_severity ON validation_findings(severity);
        CREATE INDEX IF NOT EXISTS idx_findings_code     ON validation_findings(code);
        CREATE INDEX IF NOT EXISTS idx_findings_session  ON validation_findings(session_id);
        CREATE VIEW IF NOT EXISTS v_validation_summary AS
            SELECT severity, code, COUNT(*) AS n FROM validation_findings
            GROUP BY severity, code
            ORDER BY CASE severity WHEN 'error' THEN 0 WHEN 'warning' THEN 1
                     ELSE 2 END, code;
    """)
    # Recreate the frames table if its grammar CHECK predates 'asiair_dslr'.
    frow = cur.execute(
        "SELECT sql FROM sqlite_master " "WHERE type='table' AND name='frames'"
    ).fetchone()
    if frow and "asiair_dslr" not in frow[0]:
        cur.executescript(FRAMES_TABLE_REBUILD)

    # Integrations layer. Create tables (fresh DBs), migrate columns in place
    # (existing DBs), then (re)create the views so they pick up new columns.
    cur.executescript("""
        CREATE TABLE IF NOT EXISTS integrations (
            integration_id   INTEGER PRIMARY KEY AUTOINCREMENT,
            target_id        TEXT NOT NULL REFERENCES targets(target_id),
            library_id       TEXT REFERENCES libraries(library_id),
            kind             TEXT NOT NULL CHECK (kind IN ('multi-session','composite')),
            folder_name      TEXT NOT NULL,
            folder_path      TEXT NOT NULL,
            scope            TEXT,
            sensor           TEXT,
            span             TEXT,
            version          INTEGER NOT NULL DEFAULT 1,
            session_count    INTEGER NOT NULL DEFAULT 0,
            membership_mode  TEXT NOT NULL DEFAULT 'auto',
            goal_hours       REAL,
            integration_method TEXT,
            stage_integrate  INTEGER NOT NULL DEFAULT 2,
            stage_edit       INTEGER NOT NULL DEFAULT 0,
            stage_publish    INTEGER NOT NULL DEFAULT 0,
            stage_print      INTEGER NOT NULL DEFAULT 0,
            results_file_count INTEGER NOT NULL DEFAULT 0,
            astrobin_url     TEXT,
            notes            TEXT,
            created_at       DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at       DATETIME DEFAULT CURRENT_TIMESTAMP,
            UNIQUE (folder_path)
        );
        CREATE INDEX IF NOT EXISTS idx_integrations_target ON integrations(target_id);
        CREATE INDEX IF NOT EXISTS idx_integrations_kind   ON integrations(kind);
        CREATE TABLE IF NOT EXISTS integration_members (
            integration_id INTEGER NOT NULL REFERENCES integrations(integration_id) ON DELETE CASCADE,
            session_id     INTEGER NOT NULL REFERENCES sessions(session_id) ON DELETE CASCADE,
            in_build       INTEGER NOT NULL DEFAULT 0,
            PRIMARY KEY (integration_id, session_id)
        );
        CREATE TABLE IF NOT EXISTS processing_todos (
            session_id  INTEGER NOT NULL REFERENCES sessions(session_id) ON DELETE CASCADE,
            seq         INTEGER NOT NULL,
            todo        TEXT NOT NULL,
            PRIMARY KEY (session_id, seq)
        );
    """)
    pcols = {r[1] for r in cur.execute("PRAGMA table_info(publications)")}
    if pcols and "integration_id" not in pcols:
        cur.execute(
            "ALTER TABLE publications ADD COLUMN integration_id INTEGER "
            "REFERENCES integrations(integration_id)"
        )
    # In-place column adds for a pre-existing integrations table / members table.
    icols = {r[1] for r in cur.execute("PRAGMA table_info(integrations)")}
    for col, decl in INTEGRATION_NEW_COLS:
        if col not in icols:
            cur.execute(f"ALTER TABLE integrations ADD COLUMN {col} {decl}")
    imcols = {r[1] for r in cur.execute("PRAGMA table_info(integration_members)")}
    if "in_build" not in imcols:
        cur.execute(
            "ALTER TABLE integration_members " "ADD COLUMN in_build INTEGER NOT NULL DEFAULT 0"
        )

    cur.executescript("""
        DROP VIEW IF EXISTS v_session_pipeline;
        CREATE VIEW v_session_pipeline AS
        SELECT
            s.session_id, s.target_id, t.common_name,
            s.scope, s.sensor, s.session_date, s.library_id, s.is_other_capture,
            s.stage_capture, s.stage_culled,
            s.stage_integrate, s.stage_edit, s.stage_publish, s.stage_print,
            s.integration_method,
            CASE
              WHEN s.stage_print     = 2 THEN '6 Printed'
              WHEN s.stage_publish   = 2 THEN '5 Published'
              WHEN s.stage_edit      = 2 THEN '4 Edited'
              WHEN s.stage_integrate = 2 THEN '3 Integrated'
              WHEN s.stage_culled    = 2 THEN '2 Culled'
              WHEN s.stage_capture   = 1 THEN '1 Captured'
              ELSE '0 Planned'
            END AS furthest_stage
        FROM sessions s JOIN targets t USING (target_id);

        DROP VIEW IF EXISTS v_integration_overview;
        CREATE VIEW v_integration_overview AS
        SELECT
            i.integration_id, i.target_id, t.common_name,
            i.kind, i.folder_name, i.scope, i.sensor, i.span, i.version,
            i.library_id, i.membership_mode, i.goal_hours, i.integration_method,
            COUNT(im.session_id)                                    AS sessions_available,
            SUM(COALESCE(im.in_build, 0))                           AS sessions_built,
            ROUND(SUM(s.integration_s) / 3600.0, 2)                AS available_hours,
            ROUND(SUM(CASE WHEN im.in_build = 1 THEN s.integration_s ELSE 0 END) / 3600.0, 2)
                                                                    AS built_hours,
            MAX(CASE WHEN im.in_build = 1 THEN s.session_date END)  AS data_through,
            CASE WHEN SUM(CASE WHEN im.in_build = 1 THEN 0 ELSE 1 END) > 0 THEN 1 ELSE 0 END
                                                                    AS is_stale,
            CASE
              WHEN i.stage_print   = 2 THEN '6 Printed'
              WHEN i.stage_publish = 2 THEN '5 Published'
              WHEN i.stage_edit    = 2 THEN '4 Edited'
              ELSE '3 Integrated'
            END AS furthest_stage
        FROM integrations i
        JOIN targets t USING (target_id)
        LEFT JOIN integration_members im ON im.integration_id = i.integration_id
        LEFT JOIN sessions s             ON s.session_id = im.session_id
        GROUP BY i.integration_id;

        DROP VIEW IF EXISTS v_targets_unpublished;
        CREATE VIEW v_targets_unpublished AS
        SELECT t.target_id, t.common_name, t.folder_name
        FROM targets t
        WHERE t.is_other_capture = 0
          AND NOT EXISTS (SELECT 1 FROM sessions s
                          WHERE s.target_id = t.target_id AND s.stage_publish = 2)
          AND NOT EXISTS (SELECT 1 FROM integrations i
                          WHERE i.target_id = t.target_id AND i.stage_publish = 2);

        DROP VIEW IF EXISTS v_integration_prune;
        CREATE VIEW v_integration_prune AS
        SELECT
            target_id, scope, sensor, span,
            COUNT(*)     AS version_count,
            MAX(version) AS latest_version,
            GROUP_CONCAT(folder_name, ' | ') AS folders
        FROM integrations
        WHERE kind = 'multi-session'
        GROUP BY target_id, scope, sensor, span
        HAVING COUNT(*) > 1;

        -- Calibration needs (kept in lockstep with schema.sql — recreated
        -- here so status wording changes reach existing DBs).
        DROP VIEW IF EXISTS v_calibration_needs;
        CREATE VIEW v_calibration_needs AS
        WITH raw_rollup AS (
            SELECT class, camera, scope, temperature_c, gain, exp_s,
                   COUNT(*)            AS raw_sets,
                   SUM(frame_count)    AS raw_frames,
                   MAX(capture_date)   AS newest_raw
            FROM calibration_masters
            WHERE NOT is_generated_master
            GROUP BY class, camera, scope, temperature_c, gain, exp_s
        ),
        master_rollup AS (
            SELECT class, camera, scope, temperature_c, gain, exp_s,
                   MAX(COALESCE(generated_at, capture_date)) AS master_date
            FROM calibration_masters
            WHERE is_generated_master
            GROUP BY class, camera, scope, temperature_c, gain, exp_s
        ),
        resolved AS (
            SELECT
                r.class, r.camera, r.scope, r.temperature_c, r.gain, r.exp_s,
                r.raw_sets, r.raw_frames, r.newest_raw,
                m.master_date,
                COALESCE(
                  (SELECT min_frames FROM calibration_thresholds t
                   WHERE t.class=r.class AND t.camera IS r.camera
                     AND t.temperature_c IS r.temperature_c AND t.gain IS r.gain AND t.exp_s IS r.exp_s),
                  (SELECT min_frames FROM calibration_thresholds t
                   WHERE t.class=r.class AND t.camera IS NULL AND t.scope IS NULL)
                ) AS min_frames,
                COALESCE(
                  (SELECT refresh_days FROM calibration_thresholds t
                   WHERE t.class=r.class AND t.camera IS r.camera
                     AND t.temperature_c IS r.temperature_c AND t.gain IS r.gain AND t.exp_s IS r.exp_s),
                  (SELECT refresh_days FROM calibration_thresholds t
                   WHERE t.class=r.class AND t.camera IS NULL AND t.scope IS NULL)
                ) AS refresh_days
            FROM raw_rollup r
            LEFT JOIN master_rollup m
              ON m.class=r.class AND m.camera IS r.camera AND m.scope IS r.scope
             AND m.temperature_c IS r.temperature_c AND m.gain IS r.gain AND m.exp_s IS r.exp_s
        )
        SELECT
            class, camera, scope, temperature_c, gain, exp_s,
            raw_sets, raw_frames, newest_raw, master_date, min_frames, refresh_days,
            CASE
              WHEN class='flat' THEN 'n/a (per-session)'
              WHEN master_date IS NULL THEN 'no master'
              WHEN newest_raw > master_date THEN 'stale (new raw)'
              WHEN refresh_days IS NOT NULL
                   AND julianday(date('now')) - julianday(master_date) > refresh_days
                THEN 'stale (age)'
              ELSE 'ok'
            END AS status,
            CASE WHEN raw_frames < min_frames THEN 1 ELSE 0 END AS below_threshold
        FROM resolved;

        -- Coverage recipe settings (calibration_thresholds.toml [coverage]);
        -- populated there each ingest, seeded here so the view always resolves.
        CREATE TABLE IF NOT EXISTS coverage_settings (
            id INTEGER PRIMARY KEY CHECK (id = 1),
            require_bias INTEGER NOT NULL DEFAULT 1
        );
        INSERT OR IGNORE INTO coverage_settings(id, require_bias) VALUES (1, 1);

        -- Light↔calibration coverage (full rationale in schema.sql): per
        -- (camera, gain, exposure) combo the kept lights use, is there
        -- matching dark data (±5 °C set temperature; a set with no
        -- temperature folder matches any) and bias data (per camera + gain;
        -- a set with no readable gain/ISO token matches any)? Statuses:
        -- to shoot = some subs have no matching data · to build = raws only ·
        -- stale (…) = mastered but superseded or aged (per
        -- v_calibration_needs) · ok = a fresh master covers every sub ·
        -- n/a = bias isn't in the calibration recipe (require_bias = 0).
        -- dark_low / bias_low flag matched raw sets under min_frames.
        DROP VIEW IF EXISTS v_light_calibration_coverage;
        CREATE VIEW v_light_calibration_coverage AS
        WITH lf AS (
            SELECT f.frame_id, s.sensor AS camera, f.gain, f.exp_s, f.temp_c
            FROM frames f
            JOIN sessions s USING (session_id)
            WHERE f.frame_type = 'light' AND NOT f.is_rejected
              AND f.exp_unit = 's' AND NOT s.is_other_capture
        ),
        dark_match AS (
            SELECT lf.frame_id,
                   MAX(CASE WHEN cm.is_generated_master THEN 1 ELSE 0 END) AS dk_master,
                   MAX(CASE WHEN cm.master_id IS NOT NULL THEN 1 ELSE 0 END) AS dk_any
            FROM lf
            LEFT JOIN calibration_masters cm
              ON cm.class = 'dark' AND cm.camera = lf.camera
             AND cm.gain = lf.gain AND cm.exp_s = lf.exp_s
             AND (cm.temperature_c IS NULL OR lf.temp_c IS NULL
                  OR ABS(cm.temperature_c - lf.temp_c) <= 5.0)
            GROUP BY lf.frame_id
        ),
        dark_needs AS (
            SELECT lf.frame_id,
                   MAX(CASE WHEN n.status = 'stale (new raw)' THEN 1 ELSE 0 END) AS dk_stale_raw,
                   MAX(CASE WHEN n.status = 'stale (age)' THEN 1 ELSE 0 END)     AS dk_stale_age,
                   MAX(COALESCE(n.below_threshold, 0))                           AS dk_low
            FROM lf
            LEFT JOIN v_calibration_needs n
              ON n.class = 'dark' AND n.camera = lf.camera
             AND n.gain = lf.gain AND n.exp_s = lf.exp_s
             AND (n.temperature_c IS NULL OR lf.temp_c IS NULL
                  OR ABS(n.temperature_c - lf.temp_c) <= 5.0)
            GROUP BY lf.frame_id
        ),
        bias_match AS (
            SELECT lf.frame_id,
                   MAX(CASE WHEN cm.is_generated_master THEN 1 ELSE 0 END) AS bi_master,
                   MAX(CASE WHEN cm.master_id IS NOT NULL THEN 1 ELSE 0 END) AS bi_any
            FROM lf
            LEFT JOIN calibration_masters cm
              ON cm.class = 'bias' AND cm.camera = lf.camera
             AND (cm.gain IS NULL OR cm.gain = lf.gain)
            GROUP BY lf.frame_id
        ),
        bias_needs AS (
            SELECT lf.frame_id,
                   MAX(CASE WHEN n.status = 'stale (new raw)' THEN 1 ELSE 0 END) AS bi_stale_raw,
                   MAX(CASE WHEN n.status = 'stale (age)' THEN 1 ELSE 0 END)     AS bi_stale_age,
                   MAX(COALESCE(n.below_threshold, 0))                           AS bi_low
            FROM lf
            LEFT JOIN v_calibration_needs n
              ON n.class = 'bias' AND n.camera = lf.camera
             AND (n.gain IS NULL OR n.gain = lf.gain)
            GROUP BY lf.frame_id
        )
        SELECT
            lf.camera, lf.gain, lf.exp_s,
            COUNT(*)                                AS light_subs,
            ROUND(SUM(lf.exp_s) / 3600.0, 2)        AS hours,
            MIN(lf.temp_c)                          AS temp_min,
            MAX(lf.temp_c)                          AS temp_max,
            SUM(CASE WHEN dm.dk_master THEN 1 ELSE 0 END)                   AS subs_dark_master,
            SUM(CASE WHEN dm.dk_any AND NOT dm.dk_master THEN 1 ELSE 0 END) AS subs_dark_raw,
            SUM(CASE WHEN NOT dm.dk_any THEN 1 ELSE 0 END)                  AS subs_dark_none,
            CASE
              WHEN SUM(CASE WHEN NOT dm.dk_any THEN 1 ELSE 0 END) > 0 THEN 'to shoot'
              WHEN SUM(CASE WHEN dm.dk_any AND NOT dm.dk_master THEN 1 ELSE 0 END) > 0
                THEN 'to build'
              WHEN MAX(dn.dk_stale_raw) = 1 THEN 'stale (new raw)'
              WHEN MAX(dn.dk_stale_age) = 1 THEN 'stale (age)'
              ELSE 'ok'
            END AS dark_status,
            MAX(dn.dk_low) AS dark_low,
            CASE
              WHEN (SELECT require_bias FROM coverage_settings WHERE id = 1) = 0 THEN 'n/a'
              WHEN SUM(CASE WHEN NOT bm.bi_any THEN 1 ELSE 0 END) > 0 THEN 'to shoot'
              WHEN SUM(CASE WHEN bm.bi_any AND NOT bm.bi_master THEN 1 ELSE 0 END) > 0
                THEN 'to build'
              WHEN MAX(bn.bi_stale_raw) = 1 THEN 'stale (new raw)'
              WHEN MAX(bn.bi_stale_age) = 1 THEN 'stale (age)'
              ELSE 'ok'
            END AS bias_status,
            CASE WHEN (SELECT require_bias FROM coverage_settings WHERE id = 1) = 0
                 THEN 0 ELSE MAX(bn.bi_low) END AS bias_low
        FROM lf
        JOIN dark_match dm USING (frame_id)
        JOIN dark_needs dn USING (frame_id)
        JOIN bias_match bm USING (frame_id)
        JOIN bias_needs bn USING (frame_id)
        GROUP BY lf.camera, lf.gain, lf.exp_s;
    """)
    con.commit()


def populate_vocabularies(con, org_root, log):
    """Read scope/sensor/filter/combo names from _organization/ folders."""

    def names(sub):
        p = os.path.join(org_root, sub)
        if not os.path.isdir(p):
            return []
        return sorted(
            d
            for d in os.listdir(p)
            if not d.startswith(".") and not d.startswith("!") and os.path.isdir(os.path.join(p, d))
        )

    cur = con.cursor()
    for sc in names("scope_values"):
        cur.execute(
            "INSERT INTO scopes(scope, is_imaging, from_registry) VALUES(?,0,1) "
            "ON CONFLICT(scope) DO UPDATE SET from_registry=1",
            (sc,),
        )
    for sn in names("sensor_values"):
        cur.execute(
            "INSERT INTO sensors(sensor, is_imaging, from_registry) VALUES(?,0,1) "
            "ON CONFLICT(sensor) DO UPDATE SET from_registry=1",
            (sn,),
        )
    for combo in names("scope+sensor_values"):
        if "_" in combo:
            sc, sn = combo.split("_", 1)
            # ensure both vocab rows exist before linking; mark as registry + imaging
            cur.execute(
                "INSERT INTO scopes(scope, is_imaging, from_registry) VALUES(?,1,1) "
                "ON CONFLICT(scope) DO UPDATE SET is_imaging=1, from_registry=1",
                (sc,),
            )
            cur.execute(
                "INSERT INTO sensors(sensor, is_imaging, from_registry) VALUES(?,1,1) "
                "ON CONFLICT(sensor) DO UPDATE SET is_imaging=1, from_registry=1",
                (sn,),
            )
            cur.execute(
                "INSERT OR IGNORE INTO scope_sensor_combos(scope, sensor) VALUES(?, ?)", (sc, sn)
            )
    for filt in names("filter_values"):
        # folder names carry "LABEL - description"; split on first " - "
        label = filt.split(" - ", 1)[0].strip()
        cur.execute(
            "INSERT OR IGNORE INTO filters(filter, description) VALUES(?, ?)", (label, filt)
        )
    con.commit()
    log(
        f"  vocabularies: {len(names('scope_values'))} scopes, "
        f"{len(names('sensor_values'))} sensors, "
        f"{len(names('scope+sensor_values'))} combos, "
        f"{len(names('filter_values'))} filters"
    )


def populate_planned_targets(con, org_root, log):
    """Create a target row for every registry `target folders/` entry.

    Planned targets (no capture sessions yet) then appear in the tracker at the
    'Planned' stage. Session-bearing targets are refreshed later by the library
    walk, so ON CONFLICT DO NOTHING here just seeds the not-yet-imaged ones.

    Args:
        con: open DB connection.
        org_root: the _organization folder path.
        log: logging callable.
    """
    reg = os.path.join(org_root, "target folders")
    if not os.path.isdir(reg):
        return
    cur = con.cursor()
    n = 0
    for name in sorted(os.listdir(reg)):
        if name.startswith((".", "_")) or not os.path.isdir(os.path.join(reg, name)):
            continue
        tp = parse_target_folder(name)
        cur.execute(
            """
            INSERT INTO targets(target_id, catalog, number, common_name,
                                folder_name, is_other_capture)
            VALUES(?,?,?,?,?,?)
            ON CONFLICT(target_id) DO NOTHING
        """,
            (tp["target_id"], tp["catalog"], tp["number"], tp["common_name"], name, tp["is_other"]),
        )
        n += 1
    con.commit()
    log(f"  planned targets: {n} registry entries")


def populate_target_goals(con, org_root, log):
    """Load `target_goals.toml` (array of [[goal]] blocks) into target_goals.

    Each block: target = "<folder name>", hours = <float>, priority = <int?>.
    The target folder name is resolved to a target_id; goals for unknown targets
    are skipped. Replaces the table each run (the file is the source of truth).

    Args:
        con: open DB connection.
        org_root: the _organization folder path.
        log: logging callable.
    """
    path = os.path.join(org_root, "target_goals.toml")
    if not os.path.isfile(path):
        return
    goals, block = [], None
    with open(path, encoding="utf-8") as fh:
        for raw in fh:
            line = raw.strip()
            if not line or line.startswith("#"):
                continue
            if line == "[[goal]]":
                block = {}
                goals.append(block)
                continue
            if line.startswith("["):
                block = None
                continue
            ms = re.match(r'(\w+)\s*=\s*"(.*?)"', line)
            mn = re.match(r"(\w+)\s*=\s*([\d.]+)", line)
            if ms and block is not None:
                block[ms.group(1)] = ms.group(2)
            elif mn and block is not None:
                block[mn.group(1)] = float(mn.group(2))
    cur = con.cursor()
    cur.execute("DELETE FROM target_goals")
    n = skipped = 0
    for g in goals:
        tname = g.get("target")
        if not tname:
            continue
        tid = parse_target_folder(tname)["target_id"]
        if not cur.execute("SELECT 1 FROM targets WHERE target_id=?", (tid,)).fetchone():
            skipped += 1
            continue
        cur.execute(
            """INSERT INTO target_goals(target_id, goal_hours, priority)
                       VALUES(?,?,?)""",
            (tid, g.get("hours"), int(g["priority"]) if "priority" in g else None),
        )
        n += 1
    con.commit()
    log(
        f"  target goals: {n} loaded" + (f", {skipped} skipped (unknown target)" if skipped else "")
    )


def populate_calibration_thresholds(con, org_root, log):
    """Load `calibration_thresholds.toml` (array of [[threshold]] blocks) into
    calibration_thresholds.

    Each block: class = "bias"|"dark"|"flat", min_frames = <int>, optional
    refresh_days / notes, and optional specificity keys (camera,
    temperature_c, gain, exp_s) that override the class default for matching
    sets. Replaces the table each run (the file is the source of truth);
    a missing file leaves the schema-seeded defaults in place.

    The same file's optional [coverage] section declares the calibration
    recipe: require_bias = false means bias isn't part of the workflow
    (matched darks + dark-flats), so coverage shows bias as n/a and the
    build-masters queue skips bias sets. Default: true.

    Args:
        con: open DB connection.
        org_root: the _organization folder path.
        log: logging callable.
    """
    cur = con.cursor()
    # settings table must exist (with defaults) even when the file is absent
    cur.execute("""CREATE TABLE IF NOT EXISTS coverage_settings (
                       id INTEGER PRIMARY KEY CHECK (id = 1),
                       require_bias INTEGER NOT NULL DEFAULT 1)""")
    cur.execute("INSERT OR IGNORE INTO coverage_settings(id, require_bias) VALUES (1, 1)")
    path = os.path.join(org_root, "calibration_thresholds.toml")
    if not os.path.isfile(path):
        con.commit()
        return
    rows, block, coverage, in_coverage = [], None, {}, False
    with open(path, encoding="utf-8") as fh:
        for raw in fh:
            line = raw.strip()
            if not line or line.startswith("#"):
                continue
            if line == "[[threshold]]":
                block, in_coverage = {}, False
                rows.append(block)
                continue
            if line == "[coverage]":
                block, in_coverage = None, True
                continue
            if line.startswith("["):
                block, in_coverage = None, False
                continue
            mb = re.match(r"(\w+)\s*=\s*(true|false)\b", line)
            ms = re.match(r'(\w+)\s*=\s*"(.*?)"', line)
            mn = re.match(r"(\w+)\s*=\s*(-?[\d.]+)", line)
            if in_coverage:
                if mb:
                    coverage[mb.group(1)] = mb.group(2) == "true"
                continue
            if ms and block is not None:
                block[ms.group(1)] = ms.group(2)
            elif mn and block is not None:
                block[mn.group(1)] = float(mn.group(2))
    if "require_bias" in coverage:
        cur.execute(
            "UPDATE coverage_settings SET require_bias=? WHERE id=1",
            (1 if coverage["require_bias"] else 0,),
        )
        log(
            f"  coverage recipe: require_bias=" f"{'true' if coverage['require_bias'] else 'false'}"
        )
    valid = [t for t in rows if t.get("class") in ("bias", "dark", "flat") and "min_frames" in t]
    if not valid:
        log("  calibration thresholds: file present but no valid blocks — kept existing")
        con.commit()
        return
    cur.execute("DELETE FROM calibration_thresholds")
    for t in valid:
        cur.execute(
            """INSERT INTO calibration_thresholds
                       (class, camera, scope, temperature_c, gain, exp_s,
                        min_frames, refresh_days, notes)
                       VALUES(?,?,?,?,?,?,?,?,?)""",
            (
                t["class"],
                t.get("camera"),
                t.get("scope"),
                t.get("temperature_c"),
                int(t["gain"]) if "gain" in t else None,
                t.get("exp_s"),
                int(t["min_frames"]),
                int(t["refresh_days"]) if "refresh_days" in t else None,
                t.get("notes"),
            ),
        )
    con.commit()
    log(f"  calibration thresholds: {len(valid)} loaded")


def ingest_library(con, library_id, root, obs, locations, log):
    """Walk one library: targets, sessions, frames.

    obs       — shared dict the walk appends structural observations to, for the
                later validate() pass (unparsed session folders, etc.).
    locations — {site: {lat,lon,bortle}} from locations.toml, for the Bortle lookup.
    """
    cur = con.cursor()
    n_targets = n_sessions = n_frames = 0

    for tname in sorted(os.listdir(root)):
        tpath = os.path.join(root, tname)
        # Skip dotfiles, leading-underscore utility folders (_organization,
        # _to_delete, _added, ...), and named skips.
        if tname.startswith((".", "_")) or tname in SKIP_TOPLEVEL or not os.path.isdir(tpath):
            continue
        tp = parse_target_folder(tname)

        # Upsert target. folder_path columns are per-library.
        rel_target = tname
        cur.execute(
            """
            INSERT INTO targets(target_id, catalog, number, common_name, folder_name, is_other_capture)
            VALUES(?,?,?,?,?,?)
            ON CONFLICT(target_id) DO UPDATE SET
              catalog=excluded.catalog, number=excluded.number,
              common_name=COALESCE(excluded.common_name, targets.common_name),
              folder_name=excluded.folder_name,
              is_other_capture=excluded.is_other_capture,
              updated_at=CURRENT_TIMESTAMP
        """,
            (
                tp["target_id"],
                tp["catalog"],
                tp["number"],
                tp["common_name"],
                tname,
                tp["is_other"],
            ),
        )
        n_targets += 1

        # Sessions inside this target
        for sname in sorted(os.listdir(tpath)):
            spath = os.path.join(tpath, sname)
            if sname.startswith(".") or sname.startswith("_") or not os.path.isdir(spath):
                continue
            if sname == "integrations":
                continue  # handled by ingest_integrations()
            sm = SESSION_RE.match(sname)
            if not sm:
                # Non-v2 folder name. Expected inside other-capture buckets;
                # under a deep-sky target it is a naming problem worth flagging.
                if not tp["is_other"]:
                    obs["unparsed_sessions"].append(os.path.join(rel_target, sname))
                continue
            scope, sensor, sdate = sm.group("scope"), sm.group("sensor"), sm.group("date")
            # ensure vocab rows exist (field data may pre-date a registry add)
            cur.execute("INSERT OR IGNORE INTO scopes(scope, is_imaging) VALUES(?, 1)", (scope,))
            cur.execute("INSERT OR IGNORE INTO sensors(sensor, is_imaging) VALUES(?, 1)", (sensor,))

            rel_session = os.path.join(rel_target, sname)
            # Upsert the session on its natural key; keep session_id stable.
            cur.execute(
                """
                INSERT INTO sessions(target_id, scope, sensor, session_date, library_id,
                                     folder_path, is_other_capture)
                VALUES(?,?,?,?,?,?,?)
                ON CONFLICT(target_id, scope, sensor, session_date) DO UPDATE SET
                  library_id=excluded.library_id,
                  folder_path=excluded.folder_path,
                  is_other_capture=excluded.is_other_capture,
                  updated_at=CURRENT_TIMESTAMP
            """,
                (tp["target_id"], scope, sensor, sdate, library_id, rel_session, tp["is_other"]),
            )
            sid = cur.execute(
                """SELECT session_id FROM sessions
                                 WHERE target_id=? AND scope=? AND sensor=? AND session_date=?""",
                (tp["target_id"], scope, sensor, sdate),
            ).fetchone()[0]
            n_sessions += 1

            # Re-ingest frames: delete then re-insert so counts stay exact.
            cur.execute("DELETE FROM frames WHERE session_id=?", (sid,))
            counts = {"light": 0, "flat": 0, "dark": 0, "dark_flat": 0, "bias": 0}
            rejected = 0
            integration_s = 0.0
            unparsed = 0  # FITS-extension files whose name did not parse
            header_src = None  # first kept light .fit/.fits — sampled for headers
            for abs_path, is_rej, m in walk_fits(spath):
                if m is None:
                    # Count only raw-capture files that failed to parse — not
                    # PixInsight processing output (debayered .xisf etc.) and
                    # not deliberate non-science captures (snapshots/previews).
                    if not in_processing_area(abs_path, spath) and not is_non_science(
                        os.path.basename(abs_path)
                    ):
                        unparsed += 1
                    continue
                kind = frame_kind(m)  # light/flat/dark/bias/darkflat
                ftype = "dark_flat" if kind == "darkflat" else kind
                unit = (safe(m, "unit", "s") or "s").lower()
                exp_value = float(m.group("exp"))
                exp_s = exp_value if unit == "s" else exp_value / 1000.0
                gain = m.groupdict().get("gain")
                gain = int(gain) if gain not in (None, "") else None
                temp = m.groupdict().get("temp")
                temp = float(temp) if temp not in (None, "") else None
                rot = m.groupdict().get("rot")
                rot = float(rot) if rot not in (None, "") else None
                hfr = safe(m, "hfr")
                rms = safe(m, "rms")
                # datetime — ASIAir 'dt' is YYYYMMDD-HHMMSS; NINA is date+time
                dt = m.groupdict().get("dt")
                if dt:
                    captured = f"{dt[:4]}-{dt[4:6]}-{dt[6:8]} {dt[9:11]}:{dt[11:13]}:{dt[13:15]}"
                else:
                    d = m.groupdict().get("date")
                    t = m.groupdict().get("time")
                    captured = f"{d} {t.replace('-', ':')}" if d and t else None
                binning = m.groupdict().get("bin")
                if binning is None:
                    bx = m.groupdict().get("binx")
                    by = m.groupdict().get("biny")
                    binning = f"{bx}x{by}" if bx else "?"
                grammar = (
                    "nina_v2"
                    if "HFR" in m.re.pattern
                    else (
                        "nina_legacy"
                        if "__" in m.re.pattern and "binx" in m.groupdict()
                        else (
                            "asiair_dslr"
                            if "ISO" in m.re.pattern
                            else "asiair_cal" if "target" not in m.groupdict() else "asiair_sci"
                        )
                    )
                )
                try:
                    fsize = os.path.getsize(abs_path)
                except OSError:
                    fsize = None
                cur.execute(
                    """
                    INSERT OR IGNORE INTO frames(session_id, frame_type, is_rejected,
                        exp_value, exp_unit, exp_s, binning, camera_short, gain, temp_c,
                        rotation_deg, captured_at_utc, filter, hfr, rms_arcsec,
                        grammar, file_path, file_size_bytes, sequence_index)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                    (
                        sid,
                        ftype,
                        1 if is_rej else 0,
                        exp_value,
                        unit,
                        exp_s,
                        binning,
                        m.group("cam"),
                        gain,
                        temp,
                        rot,
                        captured,
                        safe(m, "filter"),
                        float(hfr) if hfr else None,
                        float(rms) if rms else None,
                        grammar,
                        os.path.join(rel_session, os.path.relpath(abs_path, spath)),
                        fsize,
                        int(m.group("idx")) if m.groupdict().get("idx") else None,
                    ),
                )
                n_frames += 1
                if ftype == "light":
                    if is_rej:
                        rejected += 1
                    else:
                        counts["light"] += 1
                        if unit == "s":
                            integration_s += exp_s
                        if header_src is None and abs_path.lower().endswith((".fit", ".fits")):
                            header_src = abs_path
                else:
                    counts[ftype] += 1

            # Validation inputs: sample one light frame's FITS header, and read
            # the session's notes.toml. Both are best-effort.
            hdr = (
                read_fits_header(header_src)
                if header_src
                else dict(site_lat=None, site_lon=None, instrument=None, telescope=None)
            )
            notes = read_notes_toml(spath, sname)
            # A session with no FITS frames may still be a DSLR session whose
            # raws are CR3/NEF/... — count those so it is not flagged empty.
            fits_total = sum(counts.values()) + rejected
            other_images = count_raw_images(spath) if fits_total == 0 else 0
            # Files in the {session} Results folder — kept processed output.
            # A session with these is "processed", not empty, even with no raws.
            results_dir = os.path.join(spath, f"{sname} Results")
            results_files = 0
            if os.path.isdir(results_dir):
                for _r, _d, _f in os.walk(results_dir):
                    results_files += sum(
                        1 for x in _f if x != ".DS_Store" and not x.startswith("._")
                    )
            bortle = None
            if notes["location"] and notes["location"] in locations:
                bortle = locations[notes["location"]].get("bortle")
            notes_rel = (
                os.path.join(rel_session, f"{sname} notes.toml") if notes["present"] else None
            )

            # Refresh denormalised counts + validation inputs + notes metadata.
            method = detect_method(spath)  # PixInsight|PI Magic|other|None
            cur.execute(
                """
                UPDATE sessions SET
                  lights_kept=?, lights_rejected=?, flats_count=?, dark_flats_count=?,
                  darks_count=?, bias_count=?, integration_s=?,
                  unparsed_file_count=?, other_image_count=?, results_file_count=?,
                  notes_toml_present=?, stage_culled=?, stage_integrate=?,
                  stage_edit=?, stage_publish=?, stage_print=?, astrobin_url=?,
                  fits_site_lat=?, fits_site_lon=?, fits_instrument=?,
                  mount=?, location_label=?, bortle=?,
                  moon_age_days=?, moon_phase_pct=?, notes_path=?, flats_note_ref=?,
                  integration_method = CASE
                      WHEN ? IN ('PixInsight','PI Magic') THEN ?
                      ELSE COALESCE(integration_method, ?) END,
                  updated_at=CURRENT_TIMESTAMP
                WHERE session_id=?
            """,
                (
                    counts["light"],
                    rejected,
                    counts["flat"],
                    counts["dark_flat"],
                    counts["dark"],
                    counts["bias"],
                    round(integration_s, 1),
                    unparsed,
                    other_images,
                    results_files,
                    1 if notes["present"] else 0,
                    # stage_culled: rejects pulled, OR integrated (culled at stack),
                    # OR you explicitly marked a reviewed-kept-all session.
                    2 if (rejected > 0 or results_files > 0 or notes["culled"]) else 0,
                    2 if results_files > 0 else 0,  # stage_integrate: results = done
                    2 if notes["edited"] else 0,
                    2 if notes["published"] else 0,  # non-empty [[published]] list
                    2 if notes["printed"] else 0,  # non-empty [[printed]] list
                    next((e.get("url") for e in notes["published"] if e.get("url")), None),
                    hdr["site_lat"],
                    hdr["site_lon"],
                    hdr["instrument"],
                    hdr["telescope"],
                    notes["location"],
                    bortle,
                    notes["moon_age_days"],
                    notes["moon_illumination"],
                    notes_rel,
                    notes["flats_with"],
                    method,
                    method,
                    method,
                    sid,
                ),
            )
            insert_publications(
                cur, tp["target_id"], sid, None, notes["published"], notes["printed"]
            )
            cur.execute("DELETE FROM processing_todos WHERE session_id=?", (sid,))
            for i, td in enumerate(notes["todos"]):
                cur.execute(
                    "INSERT INTO processing_todos(session_id, seq, todo) " "VALUES(?,?,?)",
                    (sid, i, td),
                )

    con.commit()
    log(f"  {library_id}: {n_targets} targets, {n_sessions} sessions, {n_frames} frames")
    return n_targets, n_sessions, n_frames


def ingest_calibration(con, library_id, root, obs, log):
    """Walk _Calibration Library/ into calibration_masters.
    obs collects empty/mis-named calibration folders for the validate() pass."""
    cal_root = os.path.join(root, "_Calibration Library")
    if not os.path.isdir(cal_root):
        log(f"  {library_id}: no _Calibration Library")
        return 0
    cur = con.cursor()
    # Rebuild this library's rows from the walk — filesystem is truth, so sets
    # that were moved or renamed must not linger as orphans. Lossless: every
    # column is re-derived from disk. (Would need rethinking if the unused
    # calibration_master_inputs lineage table ever gains rows.)
    cur.execute("DELETE FROM calibration_masters WHERE library_id=?", (library_id,))
    n = 0

    def upsert(rec):
        cur.execute(
            """
            INSERT INTO calibration_masters(library_id, class, folder_path, camera, scope,
                temperature_c, gain, exp_s, capture_date, frame_count, total_size_bytes,
                is_generated_master)
            VALUES(:library_id,:class,:folder_path,:camera,:scope,:temperature_c,:gain,
                   :exp_s,:capture_date,:frame_count,:total_size_bytes,:is_generated_master)
            ON CONFLICT(folder_path) DO UPDATE SET
              camera=excluded.camera,
              scope=excluded.scope,
              temperature_c=excluded.temperature_c,
              gain=excluded.gain,
              exp_s=excluded.exp_s,
              frame_count=excluded.frame_count,
              total_size_bytes=excluded.total_size_bytes,
              capture_date=excluded.capture_date,
              is_generated_master=excluded.is_generated_master,
              updated_at=CURRENT_TIMESTAMP
        """,
            rec,
        )

    def date_of(name):
        m = DATE_RE.search(name)
        return m.group(1) if m else None

    # --- Bias: _Calibration Library/Bias/{Camera}/{Gain### | ISO###}/{Date} ---
    # (2026-07-11 layout; older flat sets directly under the camera folder
    # still parse, with gain read from frame/subfolder name tokens instead)
    bias_root = os.path.join(cal_root, "Bias")
    if os.path.isdir(bias_root):
        gain_dir_re = re.compile(r"^(?:[Gg]ain(-?\d+)|ISO(\d+))$")
        for cam in os.listdir(bias_root):
            if cam.startswith(".") or cam.startswith("!"):
                continue
            cdir = os.path.join(bias_root, cam)
            if not os.path.isdir(cdir):
                continue
            sets = []  # (set_path, set_name, gain_from_path)
            for sub in os.listdir(cdir):
                if sub.startswith(".") or "example" in sub.lower():
                    continue
                sp = os.path.join(cdir, sub)
                if not os.path.isdir(sp):
                    continue
                gm = gain_dir_re.match(sub)
                if gm:
                    gain = int(gm.group(1) or gm.group(2))
                    sets += [
                        (os.path.join(sp, d), d, gain)
                        for d in os.listdir(sp)
                        if not d.startswith(".") and os.path.isdir(os.path.join(sp, d))
                    ]
                else:
                    sets.append((sp, sub, None))
            for sp, sub, gain in sets:
                fc, sz = count_tree(sp)
                if fc == 0:
                    obs["cal_empty"].append(os.path.relpath(sp, root))
                    continue
                is_master = 1 if has_master_file(sp) else 0
                rel = os.path.relpath(sp, root)
                upsert(
                    {
                        "library_id": library_id,
                        "class": "bias",
                        "folder_path": rel,
                        "camera": cam,
                        "scope": None,
                        "temperature_c": None,
                        "gain": gain if gain is not None else detect_set_gain(sp),
                        "exp_s": None,
                        "capture_date": date_of(sub),
                        "frame_count": fc,
                        "total_size_bytes": sz,
                        "is_generated_master": is_master,
                    }
                )
                n += 1

    # --- Dark: deep tree {Camera}/{Temp}/{Gain}/{Exp}/{Dark date} ---
    dark_root = os.path.join(cal_root, "Dark")
    if os.path.isdir(dark_root):
        for camroot, dirs, files in os.walk(dark_root):
            # only act on leaf folders that actually hold frames
            real_files = [f for f in files if not f.startswith(".") and not f.startswith("._")]
            if not real_files:
                continue
            rel = os.path.relpath(camroot, root)
            parts = os.path.relpath(camroot, dark_root).split(os.sep)
            if not parts or parts[0].startswith("!"):
                continue
            camera = parts[0]
            temp = gain = exp = None
            for part in parts[1:]:
                tm = re.match(r"^(-?\d+(?:\.\d+)?)C$", part)
                gm = re.match(r"^Gain(-?\d+)$", part) or re.match(r"^ISO(\d+)$", part)
                em = re.match(r"^(\d+(?:\.\d+)?)s$", part)
                if tm:
                    temp = float(tm.group(1))
                if gm:
                    gain = int(gm.group(1))
                if em:
                    exp = float(em.group(1))
            leaf = parts[-1]
            # Some dark libraries skip the {Temp}/{Gain}/{Exp} tree and carry
            # the parameters in the ASIAir-style set name instead, e.g.
            # "Dark_300.0s_Bin1_2600MC_gain100_-20.0C". Fill whatever the path
            # segments didn't provide from the leaf name's tokens (path wins).
            if temp is None or gain is None or exp is None:
                for tok in re.split(r"[_\s]+", leaf):
                    tm = re.match(r"^(-?\d+(?:\.\d+)?)C$", tok)
                    gm = re.match(r"^[Gg]ain(-?\d+)$", tok) or re.match(r"^ISO(\d+)$", tok)
                    em = re.match(r"^(\d+(?:\.\d+)?)s$", tok)
                    if tm and temp is None:
                        temp = float(tm.group(1))
                    if gm and gain is None:
                        gain = int(gm.group(1))
                    if em and exp is None:
                        exp = float(em.group(1))
            is_master = 1 if has_master_file(camroot) else 0
            fc, sz = len(real_files), sum(
                os.path.getsize(os.path.join(camroot, f))
                for f in real_files
                if os.path.exists(os.path.join(camroot, f))
            )
            upsert(
                {
                    "library_id": library_id,
                    "class": "dark",
                    "folder_path": rel,
                    "camera": camera,
                    "scope": None,
                    "temperature_c": temp,
                    "gain": gain,
                    "exp_s": exp,
                    "capture_date": date_of(leaf),
                    "frame_count": fc,
                    "total_size_bytes": sz,
                    "is_generated_master": is_master,
                }
            )
            n += 1

    # No flat scan: flats are per-session (decided 2026-07-12, paper §11) and
    # live in session folders, never in a calibration library.

    con.commit()
    log(f"  {library_id}: {n} calibration sets")
    return n


def resolve_flats(con):
    """Derive flats_source/flats_ref for every session: where do its flats live?

    Resolution order, first match wins:
      1. 'here'          — the session folder holds flat frames.
      2. 'with sibling'  — the notes.toml [calibration] flats pointer, or a
                           same-rig same-night session that holds flats.
      3. 'none'          — no flats exist for this session.

    Flats are per-session (decided 2026-07-12, paper §11) — there is no flat
    library to fall back on. Runs after all libraries are scanned (a session and
    its flat-holding sibling can live in any library, not necessarily the same
    one) and recomputes every row, so nothing goes stale between ingests.

    Args:
        con: open SQLite connection; commits its own update.
    """
    cur = con.cursor()
    sess = cur.execute("""SELECT session_id, scope, sensor, session_date, flats_count,
                  flats_note_ref, folder_path FROM sessions""").fetchall()
    siblings = {}  # (scope, sensor, date) -> [(flats_count, folder name), ...]
    for _sid, scope, sensor, date, fc, _note, fp in sess:
        siblings.setdefault((scope, sensor, date), []).append((fc, os.path.basename(fp)))

    updates = []
    for sid, scope, sensor, date, fc, note, fp in sess:
        me = os.path.basename(fp)
        if fc > 0:
            updates.append(("here", None, sid))
        elif note:
            updates.append(("with sibling", note, sid))
        else:
            sibs = [x for x in siblings[(scope, sensor, date)] if x[0] > 0 and x[1] != me]
            if sibs:
                updates.append(("with sibling", max(sibs)[1], sid))
            else:
                updates.append(("none", None, sid))
    cur.executemany("UPDATE sessions SET flats_source=?, flats_ref=? WHERE session_id=?", updates)
    con.commit()


def ingest_integrations(con, library_id, root, obs, log):
    """Walk each target's integrations/ folder into the integrations table.

    Each subfolder of {target}/integrations/ holds an integration.toml. Members
    are resolved two ways: the AVAILABLE set (mode 'auto' → rig+span rule;
    'pinned'/legacy → the explicit `members` list) and the BUILT set (the
    [built].sessions actually stacked). Each available (or built) session becomes
    an integration_members row; in_build marks the built ones. scope/sensor and
    the multi-session-vs-composite kind are derived from the available rigs."""
    cur = con.cursor()
    n = 0

    def resolve(sname):
        """Resolve a session folder name to (session_id, scope, sensor) or None."""
        return cur.execute(
            "SELECT session_id, scope, sensor FROM sessions WHERE folder_path = ?",
            (os.path.join(tname, sname),),
        ).fetchone()

    for tname in sorted(os.listdir(root)):
        tpath = os.path.join(root, tname)
        if tname.startswith((".", "_")) or tname in SKIP_TOPLEVEL or not os.path.isdir(tpath):
            continue
        intdir = os.path.join(tpath, "integrations")
        if not os.path.isdir(intdir):
            continue
        tp = parse_target_folder(tname)
        for iname in sorted(os.listdir(intdir)):
            ipath = os.path.join(intdir, iname)
            if iname.startswith((".", "_")) or not os.path.isdir(ipath):
                continue
            rel = os.path.join(tname, "integrations", iname)
            man = read_integration_toml(os.path.join(ipath, "integration.toml"))
            if man is None:
                obs["integration_no_manifest"].append(rel)
                continue

            # Effective membership mode: explicit, else legacy list ⇒ pinned.
            if man["mode"] in ("auto", "pinned"):
                mode = man["mode"]
            elif man["members"]:
                mode = "pinned"
            else:
                mode = "auto"

            # AVAILABLE member names (what the integration should contain).
            if mode == "pinned":
                avail_names = man["members"] or man["built_sessions"]
            else:
                avail_names = resolve_auto_members(tpath, man["rig"], man["span"], man["exclude"])

            # BUILT member names (what is actually in the current master).
            if man["built_sessions"]:
                built_names = set(man["built_sessions"])
            elif mode == "pinned":
                built_names = set(avail_names)  # legacy list = the stack
            else:
                built_names = set()  # auto, not yet stacked

            # Resolve the union to session rows; track rigs of the AVAILABLE set.
            member_rows, missing, rigs = {}, [], set()
            for sname in dict.fromkeys(list(avail_names) + list(built_names)):
                row = resolve(sname)
                if not row:
                    missing.append(sname)
                    continue
                member_rows[row[0]] = sname
                if sname in avail_names:
                    rigs.add((row[1], row[2]))
            if missing:
                obs["integration_missing_member"].append((rel, missing))
            avail_set = set(avail_names)
            built_id_flags = {
                sid: (1 if sname in built_names else 0) for sid, sname in member_rows.items()
            }

            # Derive scope/sensor and kind from the available members' rigs.
            if len(rigs) == 1:
                scope, sensor = next(iter(rigs))
                derived_kind = "multi-session"
            else:
                scope, sensor = None, None
                derived_kind = "composite"
            if man["kind"] and man["kind"] != derived_kind:
                obs["integration_kind_mismatch"].append((rel, man["kind"], derived_kind))

            # Results-folder file count for this integration.
            rfiles = 0
            rdir = os.path.join(ipath, f"{iname} Results")
            if os.path.isdir(rdir):
                for _r, _d, _f in os.walk(rdir):
                    rfiles += sum(1 for x in _f if x != ".DS_Store" and not x.startswith("._"))

            sessions_available = sum(1 for s in member_rows.values() if s in avail_set)
            method = detect_method(ipath)  # PixInsight|PI Magic|other|None
            cur.execute(
                """
                INSERT INTO integrations(target_id, library_id, kind, folder_name,
                    folder_path, scope, sensor, span, version, session_count,
                    membership_mode, goal_hours, integration_method,
                    stage_integrate, stage_edit, stage_publish, stage_print,
                    results_file_count, astrobin_url)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,2,?,?,?,?,?)
                ON CONFLICT(folder_path) DO UPDATE SET
                  library_id=excluded.library_id, kind=excluded.kind,
                  folder_name=excluded.folder_name, scope=excluded.scope,
                  sensor=excluded.sensor, span=excluded.span,
                  version=excluded.version, session_count=excluded.session_count,
                  membership_mode=excluded.membership_mode,
                  goal_hours=excluded.goal_hours,
                  integration_method = CASE
                      WHEN excluded.integration_method IN ('PixInsight','PI Magic')
                        THEN excluded.integration_method
                      ELSE COALESCE(integrations.integration_method,
                                    excluded.integration_method) END,
                  stage_edit=excluded.stage_edit, stage_publish=excluded.stage_publish,
                  stage_print=excluded.stage_print,
                  results_file_count=excluded.results_file_count,
                  astrobin_url=excluded.astrobin_url, updated_at=CURRENT_TIMESTAMP
            """,
                (
                    tp["target_id"],
                    library_id,
                    derived_kind,
                    iname,
                    rel,
                    scope,
                    sensor,
                    man["span"],
                    man["version"],
                    sessions_available,
                    mode,
                    man["goal_hours"],
                    method,
                    2 if man["edited"] else 0,
                    2 if man["published"] else 0,
                    2 if man["printed"] else 0,
                    rfiles,
                    next((e.get("url") for e in man["published"] if e.get("url")), None),
                ),
            )
            iid = cur.execute(
                "SELECT integration_id FROM integrations " "WHERE folder_path=?", (rel,)
            ).fetchone()[0]
            cur.execute("DELETE FROM integration_members WHERE integration_id=?", (iid,))
            for sid, in_build in built_id_flags.items():
                cur.execute(
                    "INSERT OR IGNORE INTO integration_members"
                    "(integration_id, session_id, in_build) VALUES(?,?,?)",
                    (iid, sid, in_build),
                )
            insert_publications(cur, tp["target_id"], None, iid, man["published"], man["printed"])
            n += 1
    con.commit()
    log(f"  {library_id}: {n} integrations")
    return n


def ingest_legacy_xlsx(con, xlsx_path, log):
    """Best-effort ingest of the old Astro Acquisitions sheet into legacy_xlsx_rows."""
    if not xlsx_path or not os.path.exists(xlsx_path):
        log("  legacy xlsx: not provided / not found — skipped")
        return 0
    try:
        import openpyxl
    except ImportError:
        log("  legacy xlsx: openpyxl not installed — skipped")
        return 0
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        log(f"  legacy xlsx: could not open ({e}) — skipped")
        return 0
    if "Astro Acquisitions" not in wb.sheetnames:
        log("  legacy xlsx: no 'Astro Acquisitions' sheet — skipped")
        return 0
    ws = wb["Astro Acquisitions"]
    headers = [c.value for c in ws[1]]
    cur = con.cursor()
    n = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        as_index = str(row[0])
        data = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        cur.execute(
            """INSERT OR REPLACE INTO legacy_xlsx_rows(as_index, sheet_row_data)
                       VALUES(?, ?)""",
            (as_index, json.dumps(data, default=str)),
        )
        n += 1
    con.commit()
    log(f"  legacy xlsx: {n} rows ingested into legacy_xlsx_rows")
    return n


# ==========================================================================
# Validation pass
# ==========================================================================
def validate(con, locations, obs, log):
    """Run every data-quality check and rebuild validation_findings.

    Pure observation — never modifies the libraries. Runs at the end of each
    ingest, and standalone via validate.py. `obs` carries structural notes the
    ingest walk collected (things with no row in the DB)."""
    cur = con.cursor()
    cur.execute("DELETE FROM validation_findings")
    f = []  # (severity, code, scope, session_id, ref_path, message)

    def add(sev, code, scope, sid, ref, msg):
        f.append((sev, code, scope, sid, ref, msg))

    today = datetime.date.today().isoformat()

    # -------------------------------------------------------------- Tier 1 --
    # Session completeness (deep-sky sessions only — other-capture buckets
    # legitimately hold utility folders).
    for sid, fp, lk, lr, fl, df, dk, bi, other, unp, results in cur.execute("""
            SELECT session_id, folder_path, lights_kept, lights_rejected,
                   flats_count, dark_flats_count, darks_count, bias_count,
                   other_image_count, unparsed_file_count, results_file_count
            FROM sessions WHERE NOT is_other_capture"""):
        total = lk + lr + fl + df + dk + bi
        if results > 0:
            pass  # processed session: raws may be cleared but kept output
            # sits in the Results folder — a valid completed session
        elif total == 0 and other == 0 and unp == 0:
            add(
                "error",
                "EMPTY_SESSION",
                "session",
                sid,
                fp,
                "Session folder contains no image frames and no Results output.",
            )
        elif total == 0:
            pass  # has DSLR raws or unparsed files — covered by UNPARSED_FITS
        elif lk + lr == 0:
            add(
                "warning",
                "EMPTY_LIGHTS",
                "session",
                sid,
                fp,
                "Session has calibration frames but no light frames.",
            )

    # Folder date vs frame capture dates; future dates; multi-night span.
    for sid, fp, sdate in cur.execute("SELECT session_id, folder_path, session_date FROM sessions"):
        if sdate > today:
            add(
                "error",
                "FUTURE_DATE",
                "session",
                sid,
                fp,
                f"Session date {sdate} is in the future.",
            )
        dmin, dmax, nlight, ndays = cur.execute(
            """
            SELECT MIN(date(captured_at_utc)), MAX(date(captured_at_utc)),
                   COUNT(*), COUNT(DISTINCT date(captured_at_utc))
            FROM frames
            WHERE session_id=? AND frame_type='light'
              AND captured_at_utc IS NOT NULL""",
            (sid,),
        ).fetchone()
        if not nlight:
            continue
        nxt = (datetime.date.fromisoformat(sdate) + datetime.timedelta(days=1)).isoformat()
        on_date = cur.execute(
            """SELECT COUNT(*) FROM frames
            WHERE session_id=? AND frame_type='light'
              AND date(captured_at_utc) IN (?, ?)""",
            (sid, sdate, nxt),
        ).fetchone()[0]
        if on_date == 0:
            span = dmin if dmin == dmax else f"{dmin}..{dmax}"
            add(
                "error",
                "DATE_MISMATCH",
                "session",
                sid,
                fp,
                f"Folder dated {sdate}, but its light frames were captured {span}.",
            )
        if ndays > 2:
            add(
                "warning",
                "MULTI_NIGHT_SPAN",
                "session",
                sid,
                fp,
                f"Light frames span {ndays} calendar dates ({dmin}..{dmax}) — "
                f"may include frames copied from other sessions.",
            )

    # Unparsed FITS files inside an otherwise-ingested session.
    for sid, fp, n in cur.execute(
        "SELECT session_id, folder_path, unparsed_file_count "
        "FROM sessions WHERE unparsed_file_count > 0"
    ):
        add(
            "warning",
            "UNPARSED_FITS",
            "session",
            sid,
            fp,
            f"{n} FITS-extension file(s) here did not match any known filename "
            f"grammar — not counted in any total.",
        )

    # Scope / sensor not in the controlled vocabulary.
    for sid, fp, scope, sensor in cur.execute(
        "SELECT session_id, folder_path, scope, sensor " "FROM sessions WHERE NOT is_other_capture"
    ):
        r = cur.execute("SELECT from_registry FROM scopes WHERE scope=?", (scope,)).fetchone()
        if r and not r[0]:
            add(
                "warning",
                "UNKNOWN_SCOPE",
                "session",
                sid,
                fp,
                f"Scope '{scope}' is not in _organization/scope_values.",
            )
        r = cur.execute("SELECT from_registry FROM sensors WHERE sensor=?", (sensor,)).fetchone()
        if r and not r[0]:
            add(
                "warning",
                "UNKNOWN_SENSOR",
                "session",
                sid,
                fp,
                f"Sensor '{sensor}' is not in _organization/sensor_values.",
            )

    # Missing per-session notes.toml.
    for sid, fp in cur.execute(
        "SELECT session_id, folder_path FROM sessions "
        "WHERE NOT is_other_capture AND notes_toml_present = 0"
    ):
        add(
            "info", "NOTES_MISSING", "session", sid, fp, "No per-session notes.toml in this folder."
        )

    # notes.toml location not defined in locations.toml.
    for sid, fp, loc in cur.execute(
        "SELECT session_id, folder_path, location_label "
        "FROM sessions WHERE location_label IS NOT NULL"
    ):
        if loc not in locations:
            add(
                "error",
                "LOCATION_UNKNOWN",
                "session",
                sid,
                fp,
                f"notes.toml location '{loc}' is not defined in locations.toml.",
            )

    # Non-v2 folder names directly under a deep-sky target.
    for rel in obs.get("unparsed_sessions", []):
        add(
            "warning",
            "UNPARSED_SESSION_NAME",
            "session",
            None,
            rel,
            "Folder under a deep-sky target does not match the v2 session "
            "naming grammar — it was not ingested as a session.",
        )

    # -------------------------------------------------------------- Tier 2 --
    # FITS-header cross-checks (one sampled light header per session).
    # Registry names that legitimately differ from INSTRUME beyond vendor
    # noise words (keys/values are norm_cam()-style: lowercase alphanumeric).
    # Kept explicit — a substring match here would mask the truncated folder
    # tokens ('ASI2600MC' vs the Pro/Air bodies) validation exists to catch.
    INSTRUME_ALIASES = {"qhyminicam8m": "minicam8"}

    def norm_cam(s):
        # strip punctuation/case and the noise words manufacturers add to the
        # FITS INSTRUME string ("ZWO ...", "Canon EOS ...") so a compact folder
        # token like 'CanonR5' compares equal to INSTRUME 'Canon EOS R5'.
        n = re.sub(r"[^a-z0-9]", "", (s or "").lower()).replace("zwo", "").replace("eos", "")
        return INSTRUME_ALIASES.get(n, n)

    for sid, fp, sensor, instr in cur.execute(
        "SELECT session_id, folder_path, sensor, fits_instrument "
        "FROM sessions WHERE fits_instrument IS NOT NULL "
        "  AND NOT is_other_capture"
    ):
        if norm_cam(sensor) != norm_cam(instr):
            add(
                "warning",
                "SENSOR_MISMATCH",
                "session",
                sid,
                fp,
                f"Folder sensor '{sensor}' does not match FITS INSTRUME '{instr}'.",
            )

    for sid, fp, loc, lat, lon in cur.execute(
        "SELECT session_id, folder_path, location_label, fits_site_lat, fits_site_lon "
        "FROM sessions WHERE fits_site_lat IS NOT NULL "
        "  AND fits_site_lon IS NOT NULL AND location_label IS NOT NULL"
    ):
        site = locations.get(loc)
        if not site:
            continue  # LOCATION_UNKNOWN already covers an undefined location
        dist = ((lat - site["lat"]) ** 2 + (lon - site["lon"]) ** 2) ** 0.5
        if dist > 0.05:  # ~5 km
            add(
                "warning",
                "LOCATION_COORD_MISMATCH",
                "session",
                sid,
                fp,
                f"FITS coordinates ({lat:.3f}, {lon:.3f}) are ~{dist:.2f}° from "
                f"the declared location '{loc}'.",
            )

    # -------------------------------------------------------------- Tier 3 --
    # Registry & calibration consistency.
    for (fn,) in cur.execute("SELECT folder_name FROM targets WHERE catalog='NCG'"):
        add(
            "error",
            "CATALOG_TYPO",
            "target",
            None,
            fn,
            f"Target folder '{fn}' uses catalog 'NCG' — almost certainly a typo " f"for 'NGC'.",
        )

    reg_dir = astro_config.org_path("target folders")
    if os.path.isdir(reg_dir):
        registry = {
            d
            for d in os.listdir(reg_dir)
            if not d.startswith((".", "!")) and os.path.isdir(os.path.join(reg_dir, d))
        }
        main_folders = {r[0] for r in cur.execute("SELECT DISTINCT folder_name FROM targets")}
        for fn in sorted(main_folders - registry):
            add(
                "warning",
                "REGISTRY_MISSING",
                "registry",
                None,
                fn,
                f"Target folder '{fn}' has no entry in " f"_organization/target folders/.",
            )
        for fn in sorted(registry - main_folders):
            add(
                "info",
                "REGISTRY_ORPHAN",
                "registry",
                None,
                fn,
                f"Registry entry '{fn}' has no matching target folder in the " f"libraries.",
            )

    for rel in obs.get("cal_empty", []):
        add("info", "CAL_EMPTY", "calibration", None, rel, "Calibration folder contains no frames.")

    # Calibration folders vs the registry. Sessions are checked above
    # (UNKNOWN_SCOPE/UNKNOWN_SENSOR); a mistyped Bias/Dark camera folder
    # would otherwise slip through — and a camera folder that isn't the
    # registry sensor name can never match any light in the coverage report.
    def in_registry(table, col, name):
        r = cur.execute(f"SELECT from_registry FROM {table} WHERE {col}=?", (name,)).fetchone()
        return bool(r and r[0])

    for cls, cam in cur.execute(
        "SELECT DISTINCT class, camera FROM calibration_masters "
        "WHERE class IN ('bias','dark') AND camera IS NOT NULL"
    ):
        if not in_registry("sensors", "sensor", cam):
            add(
                "warning",
                "CAL_UNKNOWN_CAMERA",
                "calibration",
                None,
                f"_Calibration Library/{cls.capitalize()}/{cam}",
                f"Calibration camera folder '{cam}' is not in " f"_organization/sensor_values.",
            )

    # integrations — structural (from the walk)
    for rel in obs.get("integration_no_manifest", []):
        add(
            "warning",
            "INTEGRATION_NO_MANIFEST",
            "integration",
            None,
            rel,
            "Integration folder has no integration.toml manifest — not tracked.",
        )
    for rel, missing in obs.get("integration_missing_member", []):
        add(
            "error",
            "INTEGRATION_MISSING_MEMBER",
            "integration",
            None,
            rel,
            f"integration.toml lists member session(s) that don't exist: " f"{', '.join(missing)}.",
        )
    for rel, declared, derived in obs.get("integration_kind_mismatch", []):
        add(
            "warning",
            "INTEGRATION_KIND_MISMATCH",
            "integration",
            None,
            rel,
            f"integration.toml says kind='{declared}' but the members are " f"'{derived}'.",
        )

    # integrations — from the DB
    for iid, fp, sc in cur.execute(
        "SELECT integration_id, folder_path, session_count FROM integrations"
    ):
        if sc == 0:
            add(
                "error",
                "INTEGRATION_EMPTY",
                "integration",
                None,
                fp,
                "Integration has no resolvable member sessions.",
            )
    for iid, fp in cur.execute(
        "SELECT integration_id, folder_path FROM integrations "
        "WHERE kind='multi-session' AND session_count=1"
    ):
        add(
            "warning",
            "INTEGRATION_SINGLE_MEMBER",
            "integration",
            None,
            fp,
            "Multi-session integration has only one member — should it be a "
            "single-session integration, or is a member missing?",
        )

    # ------------------------------------------------------------------ write
    cur.executemany(
        "INSERT INTO validation_findings(severity,code,scope,session_id,ref_path,message)"
        " VALUES(?,?,?,?,?,?)",
        f,
    )
    con.commit()
    by = {"error": 0, "warning": 0, "info": 0}
    for rec in f:
        by[rec[0]] = by.get(rec[0], 0) + 1
    log(
        f"  validation: {len(f)} findings — "
        f"{by['error']} error, {by['warning']} warning, {by['info']} info"
    )
    return by


# ==========================================================================
# Main
# ==========================================================================
def main():
    here = os.path.dirname(os.path.abspath(__file__))
    ap = argparse.ArgumentParser(description="Populate the astrophotography tracker DB.")
    ap.add_argument("--db", default=os.path.join(here, "tracker.db"))
    ap.add_argument("--schema", default=os.path.join(here, "schema.sql"))
    ap.add_argument("--xlsx", default=None, help="path to the legacy tracker xlsx")
    ap.add_argument(
        "--config", default=None, help="path to config.toml (default: next to this script)"
    )
    ap.add_argument("--no-validate", action="store_true", help="skip the data-validation pass")
    ap.add_argument("--quiet", action="store_true")
    args = ap.parse_args()

    def log(msg):
        if not args.quiet:
            print(msg)

    log(f"Tracker ingest — {datetime.datetime.now():%Y-%m-%d %H:%M}")
    libraries = astro_config.load_libraries(args.config)
    con = init_db(args.db, args.schema)
    apply_migrations(con)  # upgrade an existing DB in place, no-op if fresh

    # Record every configured library in the libraries table (mounted or not),
    # so library_id foreign keys always resolve.
    cur0 = con.cursor()
    for lib in libraries:
        cur0.execute(
            """
            INSERT INTO libraries(library_id, label, root_path, role)
            VALUES(?,?,?,?)
            ON CONFLICT(library_id) DO UPDATE SET
              label=excluded.label, root_path=excluded.root_path,
              role=excluded.role
        """,
            (lib["id"], lib["label"], lib["path"], lib["role"]),
        )
    con.commit()

    # Vocabularies + locations come from the _organization folder, which sits
    # next to these scripts — no library path needed.
    org = astro_config.ORG_DIR
    if os.path.isdir(org):
        log("Vocabularies:")
        populate_vocabularies(con, org, log)
        populate_planned_targets(con, org, log)
    locations = load_locations(astro_config.org_path("locations.toml"))

    # Structural observations the walk collects for the validate() pass.
    obs = {
        "unparsed_sessions": [],
        "cal_empty": [],
        "integration_no_manifest": [],
        "integration_missing_member": [],
        "integration_kind_mismatch": [],
    }

    for lib in libraries:
        root = lib["path"]
        if not os.path.isdir(root):
            log(f"Library '{lib['id']}': not mounted — skipped ({root})")
            continue
        log(f"Library '{lib['id']}'  ({root})")
        ingest_library(con, lib["id"], root, obs, locations, log)
        ingest_calibration(con, lib["id"], root, obs, log)
        ingest_integrations(con, lib["id"], root, obs, log)

    # All libraries scanned — a session's flats may sit with a sibling in
    # any library, so the flats-location pass runs after the loop.
    resolve_flats(con)

    if os.path.isdir(org):
        populate_target_goals(con, org, log)
        populate_calibration_thresholds(con, org, log)

    ingest_legacy_xlsx(con, args.xlsx, log)

    if not args.no_validate:
        log("Validation:")
        validate(con, locations, obs, log)

    # Summary — report DB truth (distinct rows), not per-library operation counts.
    cur = con.cursor()

    def scalar(sql):
        v = cur.execute(sql).fetchone()[0]
        return v if v is not None else 0

    n_targets = scalar("SELECT COUNT(*) FROM targets")
    n_sessions = scalar("SELECT COUNT(*) FROM sessions")
    n_other = scalar("SELECT COUNT(*) FROM sessions WHERE is_other_capture")
    n_frames = scalar("SELECT COUNT(*) FROM frames")
    n_cal = scalar("SELECT COUNT(*) FROM calibration_masters")
    n_integ = scalar("SELECT COUNT(*) FROM integrations")
    hours = scalar(
        "SELECT ROUND(SUM(integration_s)/3600.0,2) FROM sessions WHERE NOT is_other_capture"
    )
    kept = scalar("SELECT COUNT(*) FROM frames WHERE frame_type='light' AND NOT is_rejected")
    n_err = scalar("SELECT COUNT(*) FROM validation_findings WHERE severity='error'")
    n_warn = scalar("SELECT COUNT(*) FROM validation_findings WHERE severity='warning'")
    log("")
    log("=" * 60)
    log(f"  Targets (distinct):   {n_targets}")
    log(f"  Sessions:             {n_sessions}  ({n_other} other-capture)")
    log(f"  Frames:               {n_frames}  ({kept} kept lights)")
    log(f"  Calibration sets:     {n_cal}")
    log(f"  Multi-session integ.: {n_integ}")
    log(f"  Deep-sky integration: {hours} hours")
    log(f"  Validation findings:  {n_err} errors, {n_warn} warnings")
    log("=" * 60)
    log(f"Database: {args.db}")
    con.close()


if __name__ == "__main__":
    main()
