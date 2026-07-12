#!/usr/bin/env python3
"""
export_xlsx.py — generate the Excel tracker workbook from tracker.db.

Reads the SQLite database produced by ingest.py and writes a clean multi-tab
workbook: Summary, Sessions, Targets, Calibration, QC Candidates, Integrations,
Data Health — the same data the HTML dashboard shows.

The Sessions tab carries the per-session facts from the scan. The Targets and
Summary tabs use Excel formulas (SUMIF / COUNTIF) against the Sessions tab so
the rollups stay live if a cell is edited by hand.

Usage:
    python3 export_xlsx.py [--db PATH] [--out PATH]

Run after ingest.py to refresh the workbook.
"""

import os, sys, sqlite3, argparse, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FONT = "Arial"
HEADER_BG = "1F3A5F"
HEADER_FG = "FFFFFF"
BAND_BG = "F2F5F9"
TOTAL_BG = "EAEFF5"

HEADER_FONT = Font(name=FONT, bold=True, color=HEADER_FG, size=10)
CELL_FONT = Font(name=FONT, size=10)
BOLD_FONT = Font(name=FONT, bold=True, size=10)
TITLE_FONT = Font(name=FONT, bold=True, size=14, color=HEADER_BG)
thin = Side(style="thin", color="D0D7E2")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def stage_text(v):
    return {0: "", 1: "WIP", 2: "Done"}.get(v, "")


def write_sheet(ws, headers, rows, col_formats=None, freeze="A2"):
    """Write a header row + data rows with banding and formatting."""
    col_formats = col_formats or {}
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = PatternFill("solid", start_color=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = CELL_FONT
            cell.border = BORDER
            if r % 2 == 0:
                cell.fill = PatternFill("solid", start_color=BAND_BG)
            if c in col_formats:
                cell.number_format = col_formats[c]
    ws.freeze_panes = freeze
    # auto-width
    for c in range(1, len(headers) + 1):
        letter = get_column_letter(c)
        longest = len(str(headers[c - 1]))
        for r in range(2, min(len(rows) + 2, 400)):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                longest = max(longest, len(str(v)))
        ws.column_dimensions[letter].width = min(max(longest + 2, 9), 48)


def main():
    here = os.path.dirname(os.path.abspath(__file__))
    ap = argparse.ArgumentParser(description="Generate the tracker xlsx from tracker.db")
    ap.add_argument("--db", default=os.path.join(here, "tracker.db"))
    ap.add_argument(
        "--out", default=os.path.join(here, "Astrophotography tracker (generated).xlsx")
    )
    args = ap.parse_args()

    if not os.path.exists(args.db):
        sys.exit(f"Database not found: {args.db}\nRun ingest.py first.")

    con = sqlite3.connect(args.db)
    con.row_factory = sqlite3.Row
    cur = con.cursor()
    wb = Workbook()

    def has_table(name):
        return (
            cur.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' " "AND name=?", (name,)
            ).fetchone()
            is not None
        )

    has_integrations = has_table("integrations")
    has_validation = has_table("validation_findings")

    STAGE_TEXT = {
        0: "0 Planned",
        1: "1 Captured",
        2: "2 Culled",
        3: "3 Integrated",
        4: "4 Edited",
        5: "5 Published",
        6: "6 Printed",
    }

    # ---------------------------------------------------------------- Sessions
    ws = wb.active
    ws.title = "Sessions"
    # Column layout (1-indexed): A Library B TargetID C Catalog D CommonName
    # E Scope F Sensor G SessionDate H Year I Lights J Rejected K Flats
    # L FlatsSource M FlatsLocation N DarkFlats O Darks P Bias Q BiasSource
    # R BiasLocation S Integration then stages/URL/FolderPath
    s_headers = [
        "Library",
        "Target ID",
        "Catalog",
        "Name",
        "Scope",
        "Sensor",
        "Session Date",
        "Year",
        "Lights",
        "Rejected",
        "Flats",
        "Flats Source",
        "Flats Location",
        "Dark Flats",
        "Darks",
        "Bias",
        "Bias Source",
        "Bias Location",
        "Integration (hours)",
        "Stage",
        "Method",
        "Culled",
        "Other Capture",
        "Integrate",
        "Edit",
        "Publish",
        "Print",
        "AstroBin URL",
        "Folder Path",
    ]
    s_rows = []
    for r in cur.execute("""
        SELECT s.library_id, s.target_id, t.catalog, t.common_name, s.scope, s.sensor,
               s.session_date, s.lights_kept, s.lights_rejected, s.flats_count,
               s.flats_source, s.flats_ref,
               s.dark_flats_count, s.darks_count, s.bias_count,
               s.bias_source, s.bias_ref,
               ROUND(s.integration_s/3600.0, 2) AS hrs,
               vp.furthest_stage AS stage, COALESCE(s.integration_method,'') AS method,
               s.stage_culled, s.stage_integrate, s.stage_edit, s.stage_publish,
               s.stage_print, s.astrobin_url, s.is_other_capture, s.folder_path
        FROM sessions s JOIN targets t USING (target_id)
        JOIN v_session_pipeline vp ON vp.session_id = s.session_id
        ORDER BY s.session_date, s.target_id"""):
        year = int(r["session_date"][:4]) if r["session_date"] else None
        s_rows.append(
            [
                r["library_id"],
                r["target_id"],
                r["catalog"],
                r["common_name"],
                r["scope"],
                r["sensor"],
                r["session_date"],
                year,
                r["lights_kept"],
                r["lights_rejected"],
                r["flats_count"],
                r["flats_source"] or "",
                r["flats_ref"] or "",
                r["dark_flats_count"],
                r["darks_count"],
                r["bias_count"],
                r["bias_source"] or "",
                r["bias_ref"] or "",
                r["hrs"],
                r["stage"],
                r["method"],
                stage_text(r["stage_culled"]),
                "Yes" if r["is_other_capture"] else "No",
                stage_text(r["stage_integrate"]),
                stage_text(r["stage_edit"]),
                stage_text(r["stage_publish"]),
                stage_text(r["stage_print"]),
                r["astrobin_url"] or "",
                r["folder_path"],
            ]
        )
    write_sheet(ws, s_headers, s_rows, col_formats={8: "0", 19: "0.00"})
    n = len(s_rows)
    # totals row with SUM formulas (bounded to data rows 2..n+1)
    tr = n + 2
    ws.cell(row=tr, column=1, value="TOTAL").font = BOLD_FONT
    # Lights..Integration (skip the Flats/Bias Source + Location text columns)
    for col in (9, 10, 11, 14, 15, 16, 19):
        L = get_column_letter(col)
        c = ws.cell(row=tr, column=col, value=f"=SUM({L}2:{L}{n+1})")
        c.font = BOLD_FONT
        c.fill = PatternFill("solid", start_color=TOTAL_BG)
        if col == 19:
            c.number_format = "0"
    for col in range(1, len(s_headers) + 1):
        ws.cell(row=tr, column=col).fill = PatternFill("solid", start_color=TOTAL_BG)

    # ---------------------------------------------------------------- Targets
    # Pipeline progress is now per integration, not per target. These four
    # columns are derived from the target's sessions (single-session
    # integrations) and its multi-session integrations.
    wt = wb.create_sheet("Targets")
    t_headers = [
        "Target ID",
        "Catalog",
        "Name",
        "Sessions",
        "Lifetime Hours",
        "2024 Hours",
        "2025 Hours",
        "2026 Hours",
        "Lights",
        "Scopes Used",
        "First Session",
        "Last Session",
        "Integrations",
        "Sessions Published",
        "Integrations Published",
        "Furthest Stage",
        "AstroBin URL",
        "Goal Hours",
        "% Complete",
    ]
    t_rows = []
    for r in cur.execute("""
        SELECT t.target_id, t.catalog, t.common_name, t.is_other_capture,
               t.astrobin_url,
               (SELECT GROUP_CONCAT(DISTINCT scope) FROM sessions WHERE target_id=t.target_id) AS scopes,
               (SELECT MIN(session_date) FROM sessions WHERE target_id=t.target_id) AS first_s,
               (SELECT MAX(session_date) FROM sessions WHERE target_id=t.target_id) AS last_s,
               g.goal_hours
        FROM targets t
        LEFT JOIN target_goals g USING (target_id)
        ORDER BY t.is_other_capture, t.catalog, t.number"""):
        t_rows.append(r)

    # Per-target rollups for the derived pipeline columns.
    def count_map(sql):
        return {row[0]: row[1] for row in cur.execute(sql)}

    sess_pub = count_map(
        "SELECT target_id, COUNT(*) FROM sessions " "WHERE stage_publish=2 GROUP BY target_id"
    )
    integ_ct = (
        count_map("SELECT target_id, COUNT(*) FROM integrations " "GROUP BY target_id")
        if has_integrations
        else {}
    )
    integ_pub = (
        count_map(
            "SELECT target_id, COUNT(*) FROM integrations "
            "WHERE stage_publish=2 GROUP BY target_id"
        )
        if has_integrations
        else {}
    )
    furthest = {}
    for tid, sn in cur.execute(
        "SELECT target_id, MAX(CAST(substr(furthest_stage,1,1) AS INTEGER)) "
        "FROM v_session_pipeline WHERE NOT is_other_capture GROUP BY target_id"
    ):
        furthest[tid] = sn
    if has_integrations:
        for tid, sn in cur.execute(
            "SELECT target_id, MAX(CAST(substr(furthest_stage,1,1) AS INTEGER)) "
            "FROM v_integration_overview GROUP BY target_id"
        ):
            furthest[tid] = max(furthest.get(tid, 0), sn or 0)
    # write Targets with formulas referencing Sessions
    for c, h in enumerate(t_headers, 1):
        cell = wt.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = PatternFill("solid", start_color=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    for i, r in enumerate(t_rows):
        row = i + 2
        tid = r["target_id"]
        vals = {
            1: tid,
            2: r["catalog"],
            3: r["common_name"],
            4: f"=COUNTIF(Sessions!$B:$B,$A{row})",
            5: f"=SUMIF(Sessions!$B:$B,$A{row},Sessions!$O:$O)",
            6: f"=SUMIFS(Sessions!$O:$O,Sessions!$B:$B,$A{row},Sessions!$H:$H,2024)",
            7: f"=SUMIFS(Sessions!$O:$O,Sessions!$B:$B,$A{row},Sessions!$H:$H,2025)",
            8: f"=SUMIFS(Sessions!$O:$O,Sessions!$B:$B,$A{row},Sessions!$H:$H,2026)",
            9: f"=SUMIF(Sessions!$B:$B,$A{row},Sessions!$I:$I)",
            10: r["scopes"] or "",
            11: r["first_s"] or "",
            12: r["last_s"] or "",
            13: integ_ct.get(tid, 0),
            14: sess_pub.get(tid, 0),
            15: integ_pub.get(tid, 0),
            16: STAGE_TEXT.get(furthest.get(tid, 0), "0 Planned"),
            17: r["astrobin_url"] or "",
            18: r["goal_hours"],
            19: (f'=IF(AND(ISNUMBER($R{row}),$R{row}>0),$E{row}/$R{row},"")'),
        }
        for c, v in vals.items():
            cell = wt.cell(row=row, column=c, value=v)
            cell.font = CELL_FONT
            cell.border = BORDER
            if row % 2 == 0:
                cell.fill = PatternFill("solid", start_color=BAND_BG)
            if c in (5, 6, 7, 8, 18):
                cell.number_format = "0"
            if c == 19:
                cell.number_format = "0.0%"
    wt.freeze_panes = "A2"
    for c in range(1, len(t_headers) + 1):
        letter = get_column_letter(c)
        longest = len(t_headers[c - 1])
        for row in range(2, len(t_rows) + 2):
            v = wt.cell(row=row, column=c).value
            if v and not str(v).startswith("="):
                longest = max(longest, len(str(v)))
        wt.column_dimensions[letter].width = min(max(longest + 2, 10), 42)

    # ------------------------------------------------------------ Calibration
    wc = wb.create_sheet("Calibration")
    c_headers = [
        "Class",
        "Camera",
        "Scope",
        "Temperature (°C)",
        "Gain",
        "Exposure (s)",
        "Raw Sets",
        "Raw Frames",
        "Master Date",
        "Status",
        "Below Threshold",
    ]
    c_rows = []
    for r in cur.execute("""
        SELECT class, camera, scope, temperature_c, gain, exp_s,
               raw_sets, raw_frames, master_date, status, below_threshold
        FROM v_calibration_needs
        ORDER BY class, camera, scope, temperature_c, gain, exp_s"""):
        c_rows.append(
            [
                r["class"],
                r["camera"] or "",
                r["scope"] or "",
                r["temperature_c"],
                r["gain"],
                r["exp_s"],
                r["raw_sets"],
                r["raw_frames"],
                r["master_date"] or "",
                r["status"],
                "Yes" if r["below_threshold"] else "",
            ]
        )
    write_sheet(wc, c_headers, c_rows)
    # colour the Status column
    status_colors = {
        "no master": "FCE5CD",
        "stale (new raw)": "FFF2CC",
        "stale (age)": "FFF2CC",
        "ok": "D9EAD3",
    }
    for r in range(2, len(c_rows) + 2):
        st = wc.cell(row=r, column=10).value
        if st in status_colors:
            wc.cell(row=r, column=10).fill = PatternFill("solid", start_color=status_colors[st])

    # ------------------------------------------------- Light-frame coverage
    # v_light_calibration_coverage: per (camera, gain, exposure) combo the
    # kept lights use, is matching dark/bias data on hand (see schema.sql).
    wl = wb.create_sheet("Light Coverage")
    l_headers = [
        "Camera",
        "Gain",
        "Exposure (s)",
        "Light Subs",
        "Hours",
        "Temperature Min",
        "Temperature Max",
        "Subs Dark-Mastered",
        "Subs Dark-Raw",
        "Subs No Dark",
        "Dark Status",
        "Bias Status",
        "Dark Low Frames",
        "Bias Low Frames",
    ]
    l_rows = [
        [
            r["camera"],
            r["gain"],
            r["exp_s"],
            r["light_subs"],
            r["hours"],
            r["temp_min"],
            r["temp_max"],
            r["subs_dark_master"],
            r["subs_dark_raw"],
            r["subs_dark_none"],
            r["dark_status"],
            r["bias_status"],
            "Yes" if r["dark_low"] else "",
            "Yes" if r["bias_low"] else "",
        ]
        for r in cur.execute("""
                  SELECT * FROM v_light_calibration_coverage
                  ORDER BY camera, gain, exp_s""")
    ]
    write_sheet(wl, l_headers, l_rows, col_formats={5: "0"})
    cov_colors = {
        "to shoot": "FCE5CD",
        "to build": "FFF2CC",
        "stale (new raw)": "FFF2CC",
        "stale (age)": "FFF2CC",
        "ok": "D9EAD3",
        "n/a": "EFEFEF",
    }
    for r in range(2, len(l_rows) + 2):
        for col in (11, 12):
            st = wl.cell(row=r, column=col).value
            if st in cov_colors:
                wl.cell(row=r, column=col).fill = PatternFill("solid", start_color=cov_colors[st])

    # ----------------------------------------------------------- QC candidates
    wq = wb.create_sheet("QC Candidates")
    q_headers = ["Target", "Session Date", "Captured (UTC)", "HFR", "RMS arcsec", "File"]
    q_rows = [
        [
            r["target_id"],
            r["session_date"],
            r["captured_at_utc"],
            r["hfr"],
            r["rms_arcsec"],
            r["file_path"],
        ]
        for r in cur.execute("SELECT * FROM v_qc_candidates")
    ]
    if not q_rows:
        q_rows = [['(none — no NINA v2 frames exceed HFR>3 / RMS>1.5" yet)', "", "", "", "", ""]]
    write_sheet(wq, q_headers, q_rows, col_formats={4: "0.00", 5: "0.00"})

    # ---------------------------------------------------------- Integrations
    i_rows = []
    if has_integrations:
        for r in cur.execute("""
            SELECT common_name, target_id, kind, folder_name,
                   COALESCE(scope || ' ' || sensor, 'composite') AS rig, span,
                   COALESCE(built_hours, 0) AS built_hours,
                   COALESCE(available_hours, 0) AS available_hours,
                   goal_hours, sessions_built, sessions_available,
                   data_through, COALESCE(integration_method, '') AS method,
                   is_stale, furthest_stage
            FROM v_integration_overview ORDER BY target_id, folder_name"""):
            pct = round(100.0 * r["available_hours"] / r["goal_hours"]) if r["goal_hours"] else None
            state = (
                "Stale (+%d)" % (r["sessions_available"] - r["sessions_built"])
                if r["is_stale"]
                else "Current"
            )
            i_rows.append(
                [
                    r["target_id"],
                    r["kind"],
                    r["rig"],
                    r["span"],
                    r["built_hours"],
                    r["available_hours"],
                    r["goal_hours"],
                    pct,
                    r["sessions_built"],
                    r["sessions_available"],
                    r["data_through"],
                    r["method"],
                    state,
                    r["furthest_stage"],
                    r["folder_name"],
                ]
            )
    n_integ = len(i_rows)
    wi = wb.create_sheet("Integrations")
    i_headers = [
        "Target",
        "Kind",
        "Rig",
        "Span",
        "Built (hours)",
        "Available (hours)",
        "Goal (hours)",
        "Progress %",
        "Sessions Built",
        "Sessions Available",
        "Data Through",
        "Method",
        "State",
        "Stage",
        "Folder",
    ]
    write_sheet(
        wi,
        i_headers,
        i_rows or [["(no multi-session integrations yet)"] + [""] * 14],
        col_formats={5: "0.00", 6: "0.00", 7: "0.00", 8: "0", 9: "0", 10: "0"},
    )

    # ----------------------------------------------------------- Data Health
    v_rows = []
    if has_validation:
        for r in cur.execute("""
            SELECT vf.severity, vf.code, vf.scope,
                   COALESCE(s.folder_path, vf.ref_path, '') AS location, vf.message
            FROM validation_findings vf
            LEFT JOIN sessions s ON s.session_id = vf.session_id
            ORDER BY CASE vf.severity WHEN 'error' THEN 0
                     WHEN 'warning' THEN 1 ELSE 2 END, vf.code"""):
            v_rows.append([r["severity"], r["code"], r["scope"], r["location"], r["message"]])
    n_findings = sum(1 for r in v_rows if r[0] in ("error", "warning"))
    wv = wb.create_sheet("Data Health")
    write_sheet(
        wv,
        ["Severity", "Code", "Scope", "Location", "Detail"],
        v_rows or [["(no validation findings)", "", "", "", ""]],
    )
    sev_colors = {"error": "F4CCCC", "warning": "FCE5CD", "info": "D9EAD3"}
    for r in range(2, len(v_rows) + 2):
        sv = wv.cell(row=r, column=1).value
        if sv in sev_colors:
            wv.cell(row=r, column=1).fill = PatternFill("solid", start_color=sev_colors[sv])

    # --------------------------------------------------------------- Summary
    wsum = wb.create_sheet("Summary", 0)
    wsum["A1"] = "Astrophotography Tracker"
    wsum["A1"].font = TITLE_FONT
    wsum["A2"] = f"Generated {datetime.datetime.now():%Y-%m-%d %H:%M} from tracker.db"
    wsum["A2"].font = Font(name=FONT, italic=True, size=9, color="888888")
    metrics = [
        (
            "Deep-sky integration (hours)",
            f'=SUMIFS(Sessions!$O$2:$O${n+1},Sessions!$S$2:$S${n+1},"No")',
            "0",
        ),
        ("Deep-sky sessions", f'=COUNTIF(Sessions!$S$2:$S${n+1},"No")', "0"),
        ("Other-capture sessions", f'=COUNTIF(Sessions!$S$2:$S${n+1},"Yes")', "0"),
        ("Kept light frames", f"=SUM(Sessions!$I$2:$I${n+1})", "#,##0"),
        ("Rejected light frames", f"=SUM(Sessions!$J$2:$J${n+1})", "#,##0"),
        ("Distinct targets", f"=COUNTA(Targets!$A$2:$A${len(t_rows)+1})", "0"),
        ("Calibration sets tracked", f"=COUNTA(Calibration!$A$2:$A${len(c_rows)+1})", "0"),
        (
            "Calibration items needing attention",
            '=COUNTIF(Calibration!$J:$J,"no master")+COUNTIF(Calibration!$J:$J,"stale (new raw)")+COUNTIF(Calibration!$J:$J,"stale (age)")',
            "0",
        ),
        ("Multi-session integrations", n_integ, "0"),
        ("Validation findings (error + warning)", n_findings, "0"),
    ]
    row = 4
    for label, formula, fmt in metrics:
        wsum.cell(row=row, column=1, value=label).font = BOLD_FONT
        c = wsum.cell(row=row, column=2, value=formula)
        c.font = CELL_FONT
        c.number_format = fmt
        c.fill = PatternFill("solid", start_color=BAND_BG)
        row += 1
    wsum.column_dimensions["A"].width = 36
    wsum.column_dimensions["B"].width = 16

    # Sheet order follows the pipeline flow (STYLE.md): summary, then
    # planning -> captured -> integrated, with calibration + health at the end.
    sheet_order = [
        "Summary",
        "Targets",
        "Sessions",
        "Integrations",
        "Calibration",
        "Light Coverage",
        "QC Candidates",
        "Data Health",
    ]
    wb._sheets = [wb[t] for t in sheet_order]

    wb.save(args.out)
    con.close()
    print(f"Wrote {args.out}")
    print(
        f"  Sessions: {n} · Targets: {len(t_rows)} · Calibration: {len(c_rows)} "
        f"· Integrations: {n_integ} · Health findings: {len(v_rows)}"
    )


if __name__ == "__main__":
    main()
